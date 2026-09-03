# -*- coding: utf-8 -*-
"""Excel da Conciliacao de ICMS, no layout que o Victor ja usa manualmente (ver o exemplo que ele
mandou em 31/08/2026): uma aba com os saldos por CFOP (Contabilidade x Ecomet, blocos Entradas e
Saidas) e duas abas de detalhe nota a nota (Entradas e Saidas), uma linha por nota lancada pela
Contabilidade - e' o livro que serve de base pra Dime, por isso e' o lado escolhido pro detalhe
(o topo da aba de saldos ja compara os dois lados por CFOP; aqui e' so' o suporte nota a nota do
lado que vira declaracao).
"""
from __future__ import annotations

import datetime as dt
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ...models import ConcPeriodo

FONTE = "Calibri"
AZUL = "FF1F3864"
CINZA = "FFF2F2F2"
FINA = Side(style="thin", color="FFBFBFBF")
BORDA = Border(bottom=FINA)
DINHEIRO = '"R$" #,##0.00;("R$" #,##0.00);-'
PCT = '0.00%;-0.00%;-'


def _cabecalho_grupo(ws, linha, col_ini, col_fim, texto):
    ws.merge_cells(start_row=linha, start_column=col_ini, end_row=linha, end_column=col_fim)
    c = ws.cell(linha, col_ini, texto)
    c.font = Font(name=FONTE, bold=True, size=11)
    c.alignment = Alignment(horizontal="center")


def _bloco_saldos(ws, linha, cfops, dime_por_cfop, raicms_por_cfop):
    """Uma linha por CFOP do bloco (Entradas ou Saidas): Dime nas colunas C-E, RAICMS em G-I,
    diferenca de cada um dos tres campos (Valor Contabil, Base de Calculo, Imposto Creditado) em
    K-M - pedido do Victor em 31/08/2026: antes so' tinha a diferenca do Valor Contabil."""
    inicio = linha
    for cfop in cfops:
        d = dime_por_cfop.get(cfop)
        r = raicms_por_cfop.get(cfop)
        ws.cell(linha, 2, cfop)
        ws.cell(linha, 3, float(d.valor_contabil) if d else 0).number_format = DINHEIRO
        ws.cell(linha, 4, float(d.base_calculo) if d else 0).number_format = DINHEIRO
        ws.cell(linha, 5, float(d.imposto) if d else 0).number_format = DINHEIRO
        ws.cell(linha, 7, float(r.valor_contabil) if r else 0).number_format = DINHEIRO
        ws.cell(linha, 8, float(r.base_calculo) if r else 0).number_format = DINHEIRO
        ws.cell(linha, 9, float(r.imposto) if r else 0).number_format = DINHEIRO
        ws.cell(linha, 11, f"=C{linha}-G{linha}").number_format = DINHEIRO
        ws.cell(linha, 12, f"=D{linha}-H{linha}").number_format = DINHEIRO
        ws.cell(linha, 13, f"=E{linha}-I{linha}").number_format = DINHEIRO
        for col in range(2, 14):
            ws.cell(linha, col).font = Font(name=FONTE, size=11)
            ws.cell(linha, col).border = BORDA
        linha += 1
    return linha, inicio


def _aba_saldos_cfop(wb, periodo: ConcPeriodo):
    ws = wb.active
    ws.title = "Saldos por CFOP"
    ws.sheet_view.showGridLines = False
    larguras = {"A": 3, "B": 10, "C": 16, "D": 16, "E": 17, "F": 3,
                "G": 16, "H": 16, "I": 17, "J": 3, "K": 15, "L": 15, "M": 16}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w

    dime = {(s.tipo, s.cfop): s for s in periodo.saldos if s.fonte == "dime"}
    raicms = {(s.tipo, s.cfop): s for s in periodo.saldos if s.fonte == "raicms"}
    cfops_entrada = sorted({cfop for (tipo, cfop) in list(dime) + list(raicms) if tipo == "entrada"})
    cfops_saida = sorted({cfop for (tipo, cfop) in list(dime) + list(raicms) if tipo == "saida"})
    dime_entrada = {cfop: s for (tipo, cfop), s in dime.items() if tipo == "entrada"}
    raicms_entrada = {cfop: s for (tipo, cfop), s in raicms.items() if tipo == "entrada"}
    dime_saida = {cfop: s for (tipo, cfop), s in dime.items() if tipo == "saida"}
    raicms_saida = {cfop: s for (tipo, cfop), s in raicms.items() if tipo == "saida"}

    linha = 2
    _cabecalho_grupo(ws, linha, 1, 2, "Entradas")
    _cabecalho_grupo(ws, linha, 3, 5, "Contabilidade (Dime)")
    _cabecalho_grupo(ws, linha, 7, 9, "Ecomet (RAICMS)")
    _cabecalho_grupo(ws, linha, 11, 13, "Diferença (Contabilidade − Ecomet)")
    linha += 1
    for col, txt in [(2, "CFOP"), (3, "Valor Contábil"), (4, "Base de Cálculo"),
                      (5, "Imposto Creditado"), (7, "Valor Contábil"), (8, "Base de Cálculo"),
                      (9, "Imposto Creditado"), (11, "Valor Contábil"), (12, "Base de Cálculo"),
                      (13, "Imposto Creditado")]:
        c = ws.cell(linha, col, txt)
        c.font = Font(name=FONTE, bold=True, size=11)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    linha += 1
    linha, ini_ent = _bloco_saldos(ws, linha, cfops_entrada, dime_entrada, raicms_entrada)
    fim_ent = linha - 1
    if cfops_entrada:
        ws.cell(linha, 1, "TOTAL").font = Font(name=FONTE, bold=True, size=10)
        for col in (3, 4, 5, 7, 8, 9, 11, 12, 13):
            letra = get_column_letter(col)
            cel = ws.cell(linha, col, f"=SUM({letra}{ini_ent}:{letra}{fim_ent})")
            cel.number_format = DINHEIRO
            cel.font = Font(name=FONTE, bold=True, size=10)
        linha += 1
    linha += 1

    _cabecalho_grupo(ws, linha, 1, 2, "Saídas")
    linha += 1
    linha, ini_sai = _bloco_saldos(ws, linha, cfops_saida, dime_saida, raicms_saida)
    fim_sai = linha - 1
    if cfops_saida:
        ws.cell(linha, 1, "TOTAL").font = Font(name=FONTE, bold=True, size=10)
        for col in (3, 4, 5, 7, 8, 9, 11, 12, 13):
            letra = get_column_letter(col)
            cel = ws.cell(linha, col, f"=SUM({letra}{ini_sai}:{letra}{fim_sai})")
            cel.number_format = DINHEIRO
            cel.font = Font(name=FONTE, bold=True, size=10)


def _aba_notas(wb, periodo: ConcPeriodo, tipo: str, titulo: str):
    """Uma linha por nota do Livro da Contabilidade (nao do Ecomet) - e' o livro que vira Dime,
    por isso e' o lado usado como detalhe de apoio da aba de saldos."""
    ws = wb.create_sheet(titulo)
    cabecalhos = ["CFOP", "Data de Emissão", "Número da Nota", "Valor Contábil",
                  "Alíquota de ICMS", "Valor do ICMS"]
    larguras = [10, 16, 16, 18, 16, 16]
    for i, (h, w) in enumerate(zip(cabecalhos, larguras), start=1):
        c = ws.cell(1, i, h)
        c.font = Font(name=FONTE, bold=True, size=10, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26

    notas = sorted(
        (n for n in periodo.lancamentos if n.tipo == tipo and n.origem == "contabilidade"),
        key=lambda n: (n.cfop or "", n.data_documento or dt.date.min, n.numero))

    linha = 2
    for n in notas:
        ws.cell(linha, 1, n.cfop)
        if n.data_documento:
            ws.cell(linha, 2, n.data_documento).number_format = "DD/MM/YYYY"
        ws.cell(linha, 3, n.numero)
        ws.cell(linha, 4, float(n.valor_contabil)).number_format = DINHEIRO
        ws.cell(linha, 5, float(n.aliquota) / 100).number_format = PCT
        ws.cell(linha, 6, float(n.imposto)).number_format = DINHEIRO
        for col in range(1, 7):
            ws.cell(linha, col).font = Font(name=FONTE, size=10)
            ws.cell(linha, col).border = BORDA
            if linha % 2 == 0:
                ws.cell(linha, col).fill = PatternFill("solid", fgColor=CINZA)
        linha += 1
    fim = linha - 1

    if notas:
        c = ws.cell(linha, 1, "TOTAL")
        c.font = Font(name=FONTE, bold=True, size=10)
        ws.cell(linha, 3, f"=COUNTA(C2:C{fim})").font = Font(name=FONTE, bold=True, size=10)
        for col in (4, 6):
            letra = get_column_letter(col)
            cel = ws.cell(linha, col, f"=SUM({letra}2:{letra}{fim})")
            cel.number_format = DINHEIRO
            cel.font = Font(name=FONTE, bold=True, size=10)
        fim = linha

    ws.freeze_panes = "A2"
    if fim >= 2:
        ws.auto_filter.ref = f"A1:F{fim}"


def exportar_conciliacao(periodo: ConcPeriodo) -> BytesIO:
    wb = Workbook()
    _aba_saldos_cfop(wb, periodo)
    _aba_notas(wb, periodo, "entrada", "Entradas")
    _aba_notas(wb, periodo, "saida", "Saídas")

    saida = BytesIO()
    wb.save(saida)
    saida.seek(0)
    return saida
