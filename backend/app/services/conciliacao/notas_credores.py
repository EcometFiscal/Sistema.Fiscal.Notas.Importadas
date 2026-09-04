# -*- coding: utf-8 -*-
"""Cruza a planilha de Notas de Credores (Contas a Pagar) com o Livro de Entradas da
Contabilidade de uma competência já importada na Conciliação de ICMS: preenche o CFOP de
entrada de cada nota (casando por número da NF-e + valor — mesma regra usada em
reconcile.concilia_notas para parear Contabilidade x Ecomet, reaproveitada aqui via
reconcile.enriquece_cfop) e monta uma aba de conciliação por CFOP comparando o total da
planilha com o total do Livro de Entradas. Pedido pelo Victor em 04/09/2026 (planilha de
exemplo: "Novoa Planilha", colunas UF/Emissão/Nro.NFe/Série/Modelo/Razao/Valor/ICMS/Total
ICMS/Chave).

`enriquece_cfop` já existia em reconcile.py (parte do pacote de handoff original) mas nunca
tinha sido ligada a nenhuma rota — é exatamente o casamento nota a nota que este módulo
precisava, então foi reaproveitada sem alteração de lógica.
"""
from __future__ import annotations

import unicodedata
from collections import defaultdict
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ...models import ConcPeriodo
from .exportacao import AZUL, BORDA, CINZA, DINHEIRO, FONTE
from .reconcile import enriquece_cfop

CAB_NUMERO = {'nro.nfe', 'nro nfe', 'numero', 'numero da nota', 'no.nfe', 'nfe', 'no nfe', 'n nfe'}
CAB_VALOR = {'valor', 'valor contabil', 'valor da nota'}

ROTULO_SITUACAO = {
    'EXATA': 'OK — número e valor conferem',
    'NUMERO': 'Número achado, valor não confere — conferir',
    'AMBIGUA': 'Número duplicado no Livro — conferir manualmente',
    'NAO ENCONTRADA': 'Não encontrada no Livro de Entradas',
}
COR_SITUACAO = {
    'NUMERO': 'FFFFF2CC',
    'AMBIGUA': 'FFFFE0B2',
    'NAO ENCONTRADA': 'FFF8D7DA',
}


def _normaliza(s):
    if s is None:
        return ''
    return str(s).replace('\xa0', ' ').strip()


def _chave_num(s):
    s = _normaliza(s)
    return s.lstrip('0') or '0'


def _norm_cab(s):
    s = _normaliza(s).lower()
    return ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))


def _le_planilha(conteudo: bytes):
    """Lê a primeira aba da planilha enviada -> (cabeçalhos originais, notas). Cada nota traz
    'numero'/'valor' normalizados (pro casamento) e '_linha' com os valores originais das
    células (pra reescrever a planilha de saída sem perder nenhuma coluna)."""
    wb = load_workbook(BytesIO(conteudo), data_only=True)
    ws = wb.worksheets[0]
    primeira = next(ws.iter_rows(min_row=1, max_row=1))
    cabecalhos = [_normaliza(c.value) for c in primeira]
    idx_num = idx_val = None
    for i, c in enumerate(cabecalhos):
        n = _norm_cab(c)
        if idx_num is None and n in CAB_NUMERO:
            idx_num = i
        if idx_val is None and n in CAB_VALOR:
            idx_val = i
    if idx_num is None or idx_val is None:
        raise ValueError(
            'Não encontrei as colunas "Nro.NFe" e "Valor" na planilha enviada — confira se o '
            'cabeçalho está na primeira linha e usa esses nomes (ou parecidos).')

    notas = []
    for row in ws.iter_rows(min_row=2):
        valores = [c.value for c in row]
        if all(v is None for v in valores):
            continue
        numero = _normaliza(valores[idx_num]) if idx_num < len(valores) else ''
        if not numero:
            continue
        valor = valores[idx_val] if idx_val < len(valores) else None
        notas.append(dict(numero=_chave_num(numero),
                          valor=float(valor) if isinstance(valor, (int, float)) else None,
                          _linha=valores))
    return cabecalhos, notas


def gerar_planilha_conciliada(periodo: ConcPeriodo, conteudo: bytes) -> BytesIO:
    cabecalhos, notas = _le_planilha(conteudo)
    contab_rows = [
        dict(numero=l.numero, valor_contabil=float(l.valor_contabil or 0), cfop=l.cfop,
             base_calculo=float(l.base_calculo or 0), imposto=float(l.imposto or 0))
        for l in periodo.lancamentos if l.origem == 'contabilidade' and l.tipo == 'entrada']
    enriquecidas = enriquece_cfop(notas, contab_rows)

    wb = Workbook()
    _aba_notas_credores(wb, cabecalhos, enriquecidas)
    _aba_conciliacao_cfop(wb, enriquecidas, contab_rows)

    saida = BytesIO()
    wb.save(saida)
    saida.seek(0)
    return saida


def _aba_notas_credores(wb, cabecalhos, notas):
    ws = wb.active
    ws.title = "Notas de Credores"
    todas = cabecalhos + ["CFOP de Entrada", "Situação"]
    for i, h in enumerate(todas, start=1):
        c = ws.cell(1, i, h)
        c.font = Font(name=FONTE, bold=True, size=10, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.row_dimensions[1].height = 26

    ncol = len(cabecalhos) + 1
    linha = 2
    for n in notas:
        for i, v in enumerate(n['_linha'], start=1):
            ws.cell(linha, i, v)
        ws.cell(linha, ncol, n.get('cfop') or '')
        situ = n.get('origem') or 'NAO ENCONTRADA'
        cel = ws.cell(linha, ncol + 1, ROTULO_SITUACAO.get(situ, situ))
        cor = COR_SITUACAO.get(situ)
        for col in range(1, ncol + 2):
            ws.cell(linha, col).font = Font(name=FONTE, size=10)
            ws.cell(linha, col).border = BORDA
            if cor:
                ws.cell(linha, col).fill = PatternFill("solid", fgColor=cor)
            elif linha % 2 == 0:
                ws.cell(linha, col).fill = PatternFill("solid", fgColor=CINZA)
        linha += 1
    ws.freeze_panes = "A2"
    if linha > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(todas))}{linha - 1}"


def _aba_conciliacao_cfop(wb, notas, contab_rows):
    ws = wb.create_sheet("Conciliação por CFOP")
    ws.sheet_view.showGridLines = False
    cabecalhos = ["CFOP", "Valor Planilha (Credores)", "Qtd. Notas Planilha",
                  "Valor Livro de Entrada (Contabilidade)", "Qtd. Notas Livro", "Diferença", "Situação"]
    larguras = [10, 22, 16, 28, 16, 16, 28]
    for i, (h, w) in enumerate(zip(cabecalhos, larguras), start=1):
        c = ws.cell(1, i, h)
        c.font = Font(name=FONTE, bold=True, size=10, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

    plan_agg = defaultdict(lambda: dict(valor=0.0, qtd=0))
    nao_localizadas = dict(valor=0.0, qtd=0)
    for n in notas:
        cfop, valor = n.get('cfop') or '', n.get('valor') or 0.0
        if not cfop:
            nao_localizadas['valor'] += valor
            nao_localizadas['qtd'] += 1
            continue
        for c in cfop.split('/'):
            plan_agg[c]['valor'] += valor
            plan_agg[c]['qtd'] += 1

    livro_agg = defaultdict(lambda: dict(valor=0.0, qtd=0))
    for r in contab_rows:
        livro_agg[r['cfop']]['valor'] += r['valor_contabil']
        livro_agg[r['cfop']]['qtd'] += 1

    linha = inicio = 2
    for cfop in sorted(set(plan_agg) | set(livro_agg)):
        p, l = plan_agg.get(cfop, dict(valor=0.0, qtd=0)), livro_agg.get(cfop, dict(valor=0.0, qtd=0))
        ws.cell(linha, 1, cfop)
        ws.cell(linha, 2, p['valor']).number_format = DINHEIRO
        ws.cell(linha, 3, p['qtd'])
        ws.cell(linha, 4, l['valor']).number_format = DINHEIRO
        ws.cell(linha, 5, l['qtd'])
        ws.cell(linha, 6, f"=B{linha}-D{linha}").number_format = DINHEIRO
        if cfop not in plan_agg:
            situ = 'Ausente na planilha de credores'
        elif cfop not in livro_agg:
            situ = 'Ausente no Livro de Entradas'
        elif abs(p['valor'] - l['valor']) < 0.01:
            situ = 'OK'
        else:
            situ = 'Divergente'
        ws.cell(linha, 7, situ)
        for col in range(1, 8):
            ws.cell(linha, col).font = Font(name=FONTE, size=10)
            ws.cell(linha, col).border = BORDA
        linha += 1
    fim = linha - 1
    if fim >= inicio:
        ws.cell(linha, 1, "TOTAL").font = Font(name=FONTE, bold=True, size=10)
        for col in (2, 4, 6):
            letra = get_column_letter(col)
            cel = ws.cell(linha, col, f"=SUM({letra}{inicio}:{letra}{fim})")
            cel.number_format = DINHEIRO
            cel.font = Font(name=FONTE, bold=True, size=10)
        for col in (3, 5):
            letra = get_column_letter(col)
            ws.cell(linha, col, f"=SUM({letra}{inicio}:{letra}{fim})").font = Font(name=FONTE, bold=True, size=10)
        linha += 2

    if nao_localizadas['qtd']:
        ws.cell(linha, 1, "Notas da planilha não localizadas no Livro de Entradas:").font = \
            Font(name=FONTE, bold=True, size=10)
        linha += 1
        ws.cell(linha, 1, f"{nao_localizadas['qtd']} nota(s), totalizando").font = Font(name=FONTE, size=10)
        c = ws.cell(linha, 2, nao_localizadas['valor'])
        c.number_format = DINHEIRO
        c.font = Font(name=FONTE, size=10)
