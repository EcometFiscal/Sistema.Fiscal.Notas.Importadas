"""Exportacao em Excel no layout que a contabilidade ja' recebe.

Regra de ouro: quem recebe o arquivo nao pode perceber que o sistema mudou. O layout e' o mesmo;
o que muda e' que os blocos crescem conforme o mes, em vez das 15 linhas fixas que faziam o
SUBTOTAL somar faixa errada quando o bloco estourava.
"""
from __future__ import annotations

import datetime as dt
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import Excecao, Nota, NotaItem, Produto
from . import apuracao as ap
from . import estoque as est
from . import fechamento as fec

AZUL = "FF1F3864"
CINZA = "FFF2F2F2"
FONTE = "Calibri"
FINA = Side(style="thin", color="FFBFBFBF")
BORDA = Border(left=FINA, right=FINA, top=FINA, bottom=FINA)
DINHEIRO = 'R$ #,##0.00'
PESO = '#,##0.0'
PCT = '0.00%'

CABECALHO = ["DATA", "DOCUMENTO", "NOME", "PRODUTO", "VALOR CONTÁBIL", "BASE DE CÁLCULO",
             "ALÍQUOTA", "VALOR ICMS", "ALÍQUOTA PRESUMIDO", "CRÉDITO PRESUMIDO"]
ROTULO_BLOCO = {"1": "Interestadual:  ", "2": "Interestadual:  ", "3": "Interna:  "}
# Cor de cada bloco (cabecalho + lancamentos + total), igual nas vendas e na devolucao - e' o
# que deixa visualmente obvio de qual bloco cada linha e', sem precisar ler a coluna.
COR_BLOCO = {"1": "FFFAC090", "2": "FFB7DEE8", "3": "FFC3D69B"}
AMARELO = "FFFFFF00"


def _titulo(ws, linha, texto, largura=10):
    c = ws.cell(linha, 1, texto)
    c.font = Font(name=FONTE, bold=True, size=11, color="FFFFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL)
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=largura)
    return linha + 1


def _cabecalho(ws, linha):
    for i, h in enumerate(CABECALHO, start=1):
        c = ws.cell(linha, i, h)
        c.font = Font(name=FONTE, bold=True, size=9)
        c.fill = PatternFill("solid", fgColor=CINZA)
        c.border = BORDA
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[linha].height = 26
    return linha + 1


def _bloco(ws, linha, bloco: str, carga: float, lancamentos: list[dict], regra) -> int:
    cor = COR_BLOCO[bloco]
    linha_cabecalho = linha
    linha = _cabecalho(ws, linha)
    for col in range(1, 11):
        ws.cell(linha_cabecalho, col).fill = PatternFill("solid", fgColor=cor)
    c = ws.cell(linha, 1, ROTULO_BLOCO[bloco])
    c.font = Font(name=FONTE, bold=True, size=9)
    ws.cell(linha, 2, carga).number_format = '0.0'
    for col in range(1, 11):
        ws.cell(linha, col).fill = PatternFill("solid", fgColor=cor)
    linha += 1
    inicio = linha
    for l in lancamentos:
        ws.cell(linha, 1, l["data"]).number_format = "DD/MM/YYYY"
        ws.cell(linha, 2, l["numero"])
        ws.cell(linha, 3, l["parceiro"])
        ws.cell(linha, 4, l["produto"])
        ws.cell(linha, 5, l["base"]).number_format = DINHEIRO
        ws.cell(linha, 6, l["base"]).number_format = DINHEIRO
        ws.cell(linha, 7, float(regra.aliquota)).number_format = PCT
        ws.cell(linha, 8, l["icms"]).number_format = DINHEIRO
        ws.cell(linha, 9, float(regra.aliq_presumido)).number_format = PCT
        ws.cell(linha, 10, l["credito_presumido"]).number_format = DINHEIRO
        for col in range(1, 11):
            ws.cell(linha, col).border = BORDA
            ws.cell(linha, col).fill = PatternFill("solid", fgColor=cor)
        linha += 1
    fim = linha - 1
    c = ws.cell(linha, 1, "TOTAL")
    c.font = Font(name=FONTE, bold=True, size=11)
    for col in (5, 6, 8, 10):
        letra = get_column_letter(col)
        cel = ws.cell(linha, col,
                      f"=SUM({letra}{inicio}:{letra}{fim})" if lancamentos else 0)
        cel.number_format = DINHEIRO
        cel.font = Font(name=FONTE, bold=True, size=11)
    for col in range(1, 11):
        ws.cell(linha, col).border = BORDA
        ws.cell(linha, col).fill = PatternFill("solid", fgColor=cor)
    for col in range(1, 11):
        ws.cell(linha + 1, col).fill = PatternFill("solid", fgColor=cor)
    return linha + 2


def _linha_valor(ws, linha, rotulo, valor, negrito=False, formato=DINHEIRO, col_valor=7):
    c = ws.cell(linha, 1, rotulo)
    c.font = Font(name=FONTE, bold=negrito, size=11)
    v = ws.cell(linha, col_valor, valor)
    v.number_format = formato
    v.font = Font(name=FONTE, bold=negrito, size=11)
    return linha + 1


def exportar_apuracao(db: Session, competencia: str) -> BytesIO:
    dados = ap.previa(db, competencia)
    reg = fec.registro(db, competencia)
    ano, mes = (int(x) for x in competencia.split("-"))
    ini = dt.date(ano, mes, 1)

    wb = Workbook()
    ws = wb.active
    ws.title = "OPERAÇÕES SAÍDA IMPORTADO"
    larguras = [12, 12, 42, 24, 16, 16, 10, 14, 12, 16]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    linha = _titulo(ws, 1, "SAÍDAS COM BENEFÍCIO FISCAL – TTD 409, 410 OU 411")
    por_bloco = {}
    for l in dados["lancamentos"]:
        por_bloco.setdefault(l["bloco"], []).append(l)

    for bloco in ("1", "2", "3"):
        r = ap.regra(db, bloco, ini)
        linha = _bloco(ws, linha, bloco, float(r.carga_efetiva) * 100, por_bloco.get(bloco, []), r)

    linha = _titulo(ws, linha, "DEVOLUÇÃO DE VENDAS COM BENEFÍCIO FISCAL – TTD 409, 410 OU 411")
    for bloco in ("1", "2", "3"):
        r = ap.regra(db, bloco, ini)
        linha = _bloco(ws, linha, bloco, float(r.carga_efetiva) * 100,
                       por_bloco.get(f"{bloco}D", []), r)

    linha = _titulo(ws, linha, "CRÉDITOS")
    linha = _linha_valor(ws, linha, "Créditos das operações sem benefício", None)
    linha = _linha_valor(ws, linha, "Crédito Presumido", dados["credito_presumido"])
    linha = _linha_valor(ws, linha, "Devolução ICMS", dados["devolucao_icms"])
    linha += 1

    linha = _titulo(ws, linha, "DÉBITOS")
    linha = _linha_valor(ws, linha, "ICMS pelas saídas com benefício", dados["debito"])
    linha = _linha_valor(ws, linha, "Estorno de créditos (Devoluções)", dados["estorno"])
    linha += 1

    linha = _titulo(ws, linha, "FUNDO SOCIAL")
    ws.cell(linha, 5, "Base de Cálculo").font = Font(name=FONTE, bold=True, size=9)
    ws.cell(linha, 6, "Alíquota").font = Font(name=FONTE, bold=True, size=9)
    ws.cell(linha, 7, "Valor").font = Font(name=FONTE, bold=True, size=9)
    linha += 1
    cp = dados["credito_presumido"]
    base = dados["base_beneficiada"]
    for rotulo, b, aliq in [
            ("Base de cálculo das operações alcançadas pelo benefício", base, 0.004),
            ("Valor da exoneração tributária x FUMDES", cp, 0.02),
            ("Valor da exoneração tributária x FUNDOSOCIAL", cp, 0.025)]:
        ws.cell(linha, 1, rotulo)
        ws.cell(linha, 5, b).number_format = DINHEIRO
        ws.cell(linha, 6, aliq).number_format = '0.00%'
        ws.cell(linha, 7, f"=E{linha}*F{linha}").number_format = DINHEIRO
        linha += 1
    p = linha - 3
    ws.cell(linha, 1, "Diferença a maior entre a base e o valor destinado aos fundos (a-(b+c))")
    ws.cell(linha, 7, f"=G{p}-(G{p+1}+G{p+2})").number_format = DINHEIRO
    linha += 1
    ws.cell(linha, 1, "Total").font = Font(name=FONTE, bold=True, size=11)
    ws.cell(linha, 7, f"=G{p+2}+G{linha-1}").number_format = DINHEIRO
    total_vendas = linha
    linha += 2

    linha = _titulo(ws, linha, "DEVOLUÇÕES")
    est_ = dados["estorno"]
    base_dev = sum(b["base"] for b in dados["blocos"] if b["devolucao"])
    for rotulo, b, aliq in [
            ("Base de cálculo das operações alcançadas pelo benefício", base_dev, 0.004),
            ("Valor da exoneração tributária x FUMDES", est_, 0.02),
            ("Valor da exoneração tributária x FUNDOSOCIAL", est_, 0.025)]:
        ws.cell(linha, 1, rotulo)
        ws.cell(linha, 5, b).number_format = DINHEIRO
        ws.cell(linha, 6, aliq).number_format = '0.00%'
        ws.cell(linha, 7, f"=E{linha}*F{linha}").number_format = DINHEIRO
        linha += 1
    p2 = linha - 3
    ws.cell(linha, 1, "Diferença a maior entre a base e o valor destinado aos fundos (a-(b+c))")
    ws.cell(linha, 7, f"=G{p2}-(G{p2+1}+G{p2+2})").number_format = DINHEIRO
    linha += 1
    ws.cell(linha, 1, "Total").font = Font(name=FONTE, bold=True, size=11)
    ws.cell(linha, 7, f"=G{p2+2}+G{linha-1}").number_format = DINHEIRO
    total_dev = linha
    linha += 1
    ws.cell(linha, 1, "TOTAL A RECOLHER").font = Font(name=FONTE, bold=True, size=11)
    ws.cell(linha, 7, f"=G{total_vendas}-G{total_dev}").number_format = DINHEIRO
    linha += 2

    linha = _titulo(ws, linha, "APURAÇÃO DO FUNDO DE APOIO À MANUTENÇÃO E AO DESENVOLVIMENTO "
                               "DA EDUCAÇÃO SUPERIOR")
    ws.cell(linha, 5, "Base de Calc.").font = Font(name=FONTE, bold=True, size=9)
    ws.cell(linha, 6, "Alíquota").font = Font(name=FONTE, bold=True, size=9)
    ws.cell(linha, 7, "Valor").font = Font(name=FONTE, bold=True, size=9)
    linha += 1
    for rotulo, b in [("Total do crédito presumido ou estorno de débito - TTD 409, 410 ou 411", cp),
                      ("Estorno de crédito presumido TTD 409, 410 ou 411 - Devolução", est_)]:
        ws.cell(linha, 1, rotulo)
        ws.cell(linha, 5, b).number_format = DINHEIRO
        ws.cell(linha, 6, 0.02).number_format = '0.00%'
        ws.cell(linha, 7, f"=E{linha}*F{linha}").number_format = DINHEIRO
        linha += 1
    ws.cell(linha, 1, "Fundo Educação a recolher:").font = Font(name=FONTE, bold=True, size=11)
    ws.cell(linha, 7, f"=G{linha-2}-G{linha-1}").number_format = DINHEIRO
    linha += 2

    resumo = [
        ("Débito do imposto referente as notas vinculadas ao cálculo do crédito presumido",
         dados["debito"]),
        ("Estorno de créditos referente as notas vinculadas ao cálculo do crédito presumido",
         dados["estorno"]),
        ("ICMS a deduzir da planilha de ICMS normal", dados["icms_deduzir"]),
        ("Crédito presumido calculado no período", dados["credito_presumido"]),
        ("ICMS a recolher das Oper. e Prest. beneficiadas pelo Crédito Presumido",
         dados["icms_recolher"]),
        ("Fundo Social a recolher", dados["fundo_social"]),
        ("Fundo Educação a recolher", dados["fundo_educacao"]),
    ]
    for i, (rotulo, valor) in enumerate(resumo, start=1):
        ws.cell(linha, 1, i).font = Font(name=FONTE, bold=True, size=11)
        ws.cell(linha, 2, rotulo)
        c = ws.cell(linha, 7, valor)
        c.number_format = DINHEIRO
        c.font = Font(name=FONTE, bold=True, size=11)
        if i >= 5:
            ws.cell(linha, 8, "À PAGAR").font = Font(name=FONTE, bold=True, size=11, color="FFC00000")
            for col in range(1, 9):
                ws.cell(linha, col).fill = PatternFill("solid", fgColor=AMARELO)
        linha += 1

    # entradas de importacao do mes
    ws2 = wb.create_sheet("OPERAÇÕES ENTRADA IMPORTADO")
    for i, (h, w) in enumerate(zip(
            ["Emissão", "Entrada/Saída", "Número", "Série", "Modelo", "Razão Social", "CFOP",
             "Total R$ NF", "Total QTD", "Produto", "Natureza", "NCM", "Orig+CST"],
            [12, 14, 10, 8, 8, 46, 8, 16, 14, 24, 14, 12, 10]), start=1):
        c = ws2.cell(1, i, h)
        c.font = Font(name=FONTE, bold=True, size=9)
        c.fill = PatternFill("solid", fgColor=CINZA)
        ws2.column_dimensions[get_column_letter(i)].width = w
    fim_mes = dt.date(ano + (mes == 12), (mes % 12) + 1, 1) - dt.timedelta(days=1)
    linhas = db.execute(
        select(NotaItem, Nota).join(Nota, Nota.id == NotaItem.nota_id)
        .where(Nota.tipo == "E", Nota.status != "cancelada", Nota.natureza != "ACERTO",
               Nota.data_mov.between(ini, fim_mes)).order_by(Nota.data_mov, Nota.numero)).all()
    r = 2
    for item, nota in linhas:
        ws2.cell(r, 1, nota.data_emissao or nota.data_mov).number_format = "DD/MM/YYYY"
        ws2.cell(r, 2, nota.data_mov).number_format = "DD/MM/YYYY"
        ws2.cell(r, 3, nota.numero)
        ws2.cell(r, 4, nota.serie)
        ws2.cell(r, 5, nota.modelo)
        ws2.cell(r, 6, nota.parceiro.nome if nota.parceiro else None)
        ws2.cell(r, 7, nota.cfop)
        ws2.cell(r, 8, float(item.valor or 0)).number_format = DINHEIRO
        ws2.cell(r, 9, float(item.quantidade or 0)).number_format = PESO
        ws2.cell(r, 10, item.produto.descricao)
        ws2.cell(r, 11, nota.natureza)
        ws2.cell(r, 12, item.ncm)
        ws2.cell(r, 13, item.cst_completo)
        r += 1

    ws3 = wb.create_sheet("ORIGEM DO ARQUIVO")
    ws3.column_dimensions["A"].width = 34
    ws3.column_dimensions["B"].width = 70
    info = [
        ("Competência", competencia),
        ("Gerado em", dt.datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Situação da competência", (reg.status if reg else "aberta")),
        ("Fechada por", (reg.fechada_por if reg and reg.status == "fechada" else "—")),
        ("Notas na apuração", len(dados["lancamentos"])),
        ("Base beneficiada", dados["base_beneficiada"]),
        ("ICMS a recolher", dados["icms_recolher"]),
        ("Carga efetiva do mês", f"{dados['carga_efetiva']:.3f}%".replace(".", ",")),
        ("", ""),
        ("Como este arquivo foi montado",
         "Os valores vêm dos lançamentos da competência, não de células digitadas. Os totais de "
         "cada bloco são fórmulas sobre o intervalo exato do bloco, que cresce conforme o mês — "
         "não há mais faixa fixa de 15 linhas."),
        ("Conferência",
         "Recalcular esta competência em qualquer data devolve o mesmo número enquanto os "
         "lançamentos não mudarem. Se a competência estiver fechada, o sistema também guarda os "
         "totais congelados no fechamento."),
    ]
    for i, (a, b) in enumerate(info, start=1):
        ws3.cell(i, 1, a).font = Font(name=FONTE, bold=True, size=9)
        c = ws3.cell(i, 2, b)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if isinstance(b, float):
            c.number_format = DINHEIRO

    saida = BytesIO()
    wb.save(saida)
    saida.seek(0)
    return saida


def exportar_estoque(db: Session, de: dt.date | None, ate: dt.date | None) -> BytesIO:
    ate = ate or dt.date.today()
    de = de or dt.date(2020, 11, 1)
    anterior = {p["produto_id"]: p for p in est.posicao(db, de - dt.timedelta(days=1))}
    atual = est.posicao(db, ate)

    wb = Workbook()
    ws = wb.active
    ws.title = "POSIÇÃO"
    cabecalhos = ["PRODUTO", "SALDO ANTERIOR (KG)", "SALDO ANTERIOR (R$)", "ENTRADAS (KG)",
                  "SAÍDAS (KG)", "SALDO FINAL (KG)", "SALDO FINAL (R$)", "CUSTO MÉDIO (R$/KG)"]
    for i, (h, w) in enumerate(zip(cabecalhos, [26, 18, 18, 16, 16, 16, 18, 18]), start=1):
        c = ws.cell(1, i, h)
        c.font = Font(name=FONTE, bold=True, size=9, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

    linha = 2
    for p in atual:
        movs = [(i, n) for i, n in est.movimentos(db, p["produto_id"], ate)
                if n.data_mov >= de]
        entradas = sum(float(i.quantidade) for i, n in movs if n.tipo == "E")
        saidas = sum(float(i.quantidade) for i, n in movs if n.tipo == "S")
        ant = anterior.get(p["produto_id"], dict(saldo_kg=0, saldo_rs=0))
        for col, valor, fmt in [
                (1, p["produto"], None), (2, ant["saldo_kg"], PESO), (3, ant["saldo_rs"], DINHEIRO),
                (4, entradas, PESO), (5, saidas, PESO), (6, p["saldo_kg"], PESO),
                (7, p["saldo_rs"], DINHEIRO), (8, p["custo_medio"], DINHEIRO)]:
            c = ws.cell(linha, col, valor)
            if fmt:
                c.number_format = fmt
            c.border = BORDA
        linha += 1
    ws.cell(linha, 1, "TOTAL").font = Font(name=FONTE, bold=True, size=11)
    for col in (2, 3, 4, 5, 6, 7):
        letra = get_column_letter(col)
        c = ws.cell(linha, col, f"=SUM({letra}2:{letra}{linha-1})")
        c.number_format = PESO if col in (2, 4, 5, 6) else DINHEIRO
        c.font = Font(name=FONTE, bold=True, size=11)
        c.fill = PatternFill("solid", fgColor=CINZA)

    ws2 = wb.create_sheet("MOVIMENTAÇÃO")
    cab2 = ["DATA", "OPERAÇÃO", "NATUREZA", "NF", "PARCEIRO", "PRODUTO", "QTD (KG)", "VALOR (R$)",
            "CUSTO DA SAÍDA (R$)", "SALDO DO PRODUTO (KG)", "USUÁRIO"]
    for i, (h, w) in enumerate(zip(cab2, [12, 12, 14, 10, 44, 24, 14, 16, 18, 20, 14]), start=1):
        c = ws2.cell(1, i, h)
        c.font = Font(name=FONTE, bold=True, size=9, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        ws2.column_dimensions[get_column_letter(i)].width = w
    r = 2
    for produto in db.execute(select(Produto).order_by(Produto.descricao)).scalars():
        for l in est.razao(db, produto.id, de, ate):
            ws2.cell(r, 1, l["data"]).number_format = "DD/MM/YYYY"
            ws2.cell(r, 2, "ENTRADA" if l["tipo"] == "E" else "SAÍDA")
            ws2.cell(r, 3, l["natureza"])
            ws2.cell(r, 4, l["numero"] or None)
            ws2.cell(r, 5, l["parceiro"])
            ws2.cell(r, 6, produto.descricao)
            ws2.cell(r, 7, l["quantidade"]).number_format = PESO
            ws2.cell(r, 8, l["valor"]).number_format = DINHEIRO
            ws2.cell(r, 9, l["custo_total"]).number_format = DINHEIRO
            ws2.cell(r, 10, l["saldo"]).number_format = PESO
            r += 1
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:K{r-1}"

    saida = BytesIO()
    wb.save(saida)
    saida.seek(0)
    return saida


# Area de origem de cada tipo de pendencia, pro relatorio de erros agrupar o que a apuracao,
# o estoque e a importacao de dados encontraram - sem essa separacao vira uma lista solta e
# ninguem sabe se um item e' problema fiscal ou so' informativo.
AREA_POR_TIPO = {
    "importacao_xml": "Importação",
    "casamento_ambiguo": "Importação",
    "cnpj_detectado": "Importação",
    "nota_sem_data": "Importação",
    "acerto_automatico": "Estoque",
    "saida_sem_saldo": "Estoque",
    "duplicata_confirmada": "Lançamento manual",
}


def exportar_pendencias(db: Session) -> BytesIO:
    """Relatorio de erros/pendencias que o sistema encontrou em apuracao, estoque ou
    importacao de dados - tudo que foi aceito mas alguem precisa olhar (Excecao), numa aba por
    area e um resumo no topo."""
    linhas = db.execute(
        select(Excecao, Nota).outerjoin(Nota, Nota.id == Excecao.nota_id)
        .order_by(Excecao.resolvida, Excecao.criado_em.desc())).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "RESUMO"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    linha = _titulo(ws, 1, "RELATÓRIO DE ERROS E PENDÊNCIAS", largura=2)
    ws.cell(linha, 1, "Gerado em").font = Font(name=FONTE, bold=True, size=9)
    ws.cell(linha, 2, dt.datetime.now().strftime("%d/%m/%Y %H:%M"))
    linha += 1
    ws.cell(linha, 1, "Total de pendências").font = Font(name=FONTE, bold=True, size=9)
    ws.cell(linha, 2, len(linhas))
    linha += 1
    ws.cell(linha, 1, "Não resolvidas").font = Font(name=FONTE, bold=True, size=9)
    ws.cell(linha, 2, sum(1 for e, _ in linhas if not e.resolvida))
    linha += 2
    ws.cell(linha, 1, "Por área").font = Font(name=FONTE, bold=True, size=10)
    linha += 1
    ws.cell(linha, 1, "Área").font = Font(name=FONTE, bold=True, size=9)
    ws.cell(linha, 2, "Pendências").font = Font(name=FONTE, bold=True, size=9)
    linha += 1
    por_area: dict[str, int] = {}
    for e, _ in linhas:
        area = AREA_POR_TIPO.get(e.tipo, "Outros")
        por_area[area] = por_area.get(area, 0) + 1
    for area, n in sorted(por_area.items(), key=lambda x: -x[1]):
        ws.cell(linha, 1, area)
        ws.cell(linha, 2, n)
        linha += 1

    ws2 = wb.create_sheet("PENDÊNCIAS")
    cab = ["ÁREA", "TIPO", "QUANDO", "RESOLVIDA", "NF", "OPERAÇÃO", "DESCRIÇÃO", "JUSTIFICATIVA",
           "QUANTIDADE (KG)", "VALOR (R$)", "USUÁRIO"]
    for i, (h, w) in enumerate(zip(cab, [16, 20, 14, 11, 10, 10, 60, 30, 16, 16, 14]), start=1):
        c = ws2.cell(1, i, h)
        c.font = Font(name=FONTE, bold=True, size=9, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[1].height = 26
    r = 2
    for e, nota in linhas:
        ws2.cell(r, 1, AREA_POR_TIPO.get(e.tipo, "Outros"))
        ws2.cell(r, 2, e.tipo)
        ws2.cell(r, 3, e.criado_em).number_format = "DD/MM/YYYY HH:MM"
        ws2.cell(r, 4, "SIM" if e.resolvida else "NÃO")
        ws2.cell(r, 5, nota.numero if nota else None)
        ws2.cell(r, 6, ("ENTRADA" if nota.tipo == "E" else "SAÍDA") if nota else None)
        ws2.cell(r, 7, e.descricao)
        ws2.cell(r, 8, e.justificativa)
        ws2.cell(r, 9, float(e.quantidade) if e.quantidade else None).number_format = PESO
        ws2.cell(r, 10, float(e.valor) if e.valor else None).number_format = DINHEIRO
        ws2.cell(r, 11, e.criado_por)
        for col in range(1, 12):
            ws2.cell(r, col).border = BORDA
        r += 1
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:K{r-1}"

    saida = BytesIO()
    wb.save(saida)
    saida.seek(0)
    return saida
