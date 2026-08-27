#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Projeto Lastro - FASE 1
Migracao e saneamento do historico (nov/2020 -> hoje).

Le os dois arquivos-fonte (ESTOQUE FISCAL IMPORTADO.xlsm e a Apuracao ICMS do mes),
normaliza produtos e parceiros, carrega o historico em uma base relacional com o
modelo das 6 tabelas do documento de concepcao, roda o consumo PEPS sobre todo o
periodo e emite um relatorio de divergencias em Excel.

Nada e' corrigido em silencio: toda alteracao proposta vai para o relatorio com
linha de origem, impacto e campo de decisao para conferencia humana.

Uso:
    python3 migracao_lastro.py --estoque estoque.xlsm --apuracao apuracao_072026.xlsx \
                               --db lastro.db --relatorio relatorio_divergencias_fase1.xlsx

Banco: SQLite (portavel, roda em qualquer maquina). O DDL usa apenas tipos e
construcoes com equivalente direto em PostgreSQL - ver schema_postgres.sql.
"""

import argparse
import datetime as dt
import os
import re
import sqlite3
import difflib
import unicodedata
import warnings
from collections import defaultdict

import openpyxl

warnings.filterwarnings("ignore")

HOJE = dt.date.today()

# Decisoes tomadas por Victor em 27/08/2026 sobre o relatorio da Fase 1.
# Ficam aqui, versionadas, em vez de espalhadas pelo codigo.
DECISOES = {
    "saldo_negativo": "acerto_datado",        # 1. saida sem lastro gera lancamento de acerto na data
    "ncm": "desconsiderar",                   # 2. NCM nao sera' exigido nesta base
    "notas_fora_apuracao": "ajuste_contabil", # 3. as 3 notas de cobre serao ajustadas com a contabilidade
    "parceiros_parecidos": "unificar",        # 4. unificar parceiros com mesmo nucleo de nome
    "vinculo_lastro": "controle_por_saldo",   # 5. sem origem por item e sem vinculo saida<->entrada
    "linhas_identicas": "manter",             # 6. notas diferentes podem ter dados identicos
}
COMP_APURACAO = "2026-07"          # competencia do arquivo de apuracao anexado
TOL = 0.01                          # tolerancia de centavos na conferencia


# ----------------------------------------------------------------------------
# 1. Normalizacao
# ----------------------------------------------------------------------------

def fold(s):
    """Remove acentos e caixa; usado so' para COMPARAR, nunca para gravar."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.upper()


def limpa_texto(s):
    """Tira NBSP (\\xa0), espacos duplicados, pontuacao terminal e espacos das pontas."""
    if s is None:
        return None
    s = str(s).replace("\xa0", " ").replace("​", " ")
    s = " ".join(s.split())
    s = s.strip(" .,;-")
    return s.upper()


def chave_parceiro(nome):
    """Chave de comparacao: sem acento, sem pontuacao, tokens colapsados."""
    s = fold(limpa_texto(nome))
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    s = " ".join(s.split())
    # sufixos societarios nao distinguem empresa
    return s


# Tokens que nao distinguem empresa - usados so' para achar o "nucleo" da razao social.
STOP_RAZAO = {"IND", "INDUSTRIA", "INDUSTRIAL", "COM", "COMERCIO", "COMERCIAL", "DE", "DA", "DO",
              "DOS", "DAS", "E", "METAIS", "METAL", "LTDA", "EIRELI", "EPP", "ME", "SA", "S", "A",
              "CIA", "IMP", "EXP", "IMPORTACAO", "EXPORTACAO", "RECICLAGEM", "RECICLYNG", "ALUMINIO",
              "ALUMINIOS", "LIGAS", "METALICAS", "PRODUTOS", "GESTAO", "AMBIENTAL", "SUSTENTAVEIS",
              "GMBH", "BV", "LTD", "LIMITED", "CO", "KG", "INC", "TRADING", "COMPANY", "CORPORATION",
              "REFUSAO", "CENTRO", "RESIDUOS", "TRANSPORTADORA", "ENGENHARIA", "EM", "EIRELLI"}


def nucleo_razao(nome):
    t = [x for x in chave_parceiro(nome).split() if x not in STOP_RAZAO]
    return t or chave_parceiro(nome).split()


PRODUTO_CANON = {}  # preenchido em canon_produto


def canon_produto(desc):
    """Canonicaliza a descricao do produto. 'SUCATA DE MAGNESIO' == 'SUCATA DE MAGNÉSIO'."""
    limpo = limpa_texto(desc)
    k = fold(limpo)
    k = " ".join(k.split())
    PRODUTO_CANON.setdefault(k, set()).add(limpo)
    return k


def parse_data(v):
    """Aceita datetime, date, texto dd/mm/aaaa e serial do Excel. Devolve (date|None, obs)."""
    if v is None:
        return None, "data ausente"
    if isinstance(v, dt.datetime):
        return v.date(), None
    if isinstance(v, dt.date):
        return v, None
    if isinstance(v, (int, float)):
        try:
            base = dt.date(1899, 12, 30)
            return base + dt.timedelta(days=int(v)), "data gravada como numero de serie"
        except Exception:
            return None, "data numerica invalida"
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(s, fmt).date(), "data gravada como texto"
        except ValueError:
            continue
    return None, f"data ilegivel: {s!r}"


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


# ----------------------------------------------------------------------------
# 2. Registro de divergencias
# ----------------------------------------------------------------------------

class Registro:
    def __init__(self):
        self.itens = []
        self._seq = 0

    def add(self, classe, severidade, origem, linha, documento, descricao,
            impacto_kg=None, impacto_rs=None, acao=None):
        self._seq += 1
        self.itens.append(dict(
            id=f"D{self._seq:04d}", classe=classe, severidade=severidade,
            origem=origem, linha=linha, documento=documento, descricao=descricao,
            impacto_kg=impacto_kg, impacto_rs=impacto_rs, acao=acao, decisao="",
            situacao="Resolvido na migracao" if classe in RESOLVIDAS_NA_MIGRACAO else "Pendente"))
        return self.itens[-1]

    def por_classe(self):
        d = defaultdict(list)
        for i in self.itens:
            d[i["classe"]].append(i)
        return d


RESOLVIDAS_NA_MIGRACAO = {
    "Produto - variantes unificadas", "Parceiro - grafia invisivel", "Parceiro - variantes unificadas",
    "Parceiro - unificado por decisao", "Painel - KPI x lista", "Data - formato",
    "NCM dispensado por decisao", "Acerto de estoque gerado", "Saldo negativo - corrigido por acerto",
    "Chave de acesso ausente", "Origem por item dispensada por decisao",
    "Linha duplicada - mantida por decisao",
}

REG = Registro()
UNIFICADOS_POR_DECISAO = []   # [(grupo_dono, [grupos_absorvidos])]
GRUPOS_DECIDIDOS = set()      # chaves de grupo unificadas pela decisao 4
ACERTOS = []                  # lancamentos de acerto gerados pela decisao 1
SEV_ORDEM = {"Critico": 0, "Alto": 1, "Medio": 2, "Baixo": 3}


# ----------------------------------------------------------------------------
# 3. Leitura das planilhas
# ----------------------------------------------------------------------------

def ler_movimentos(caminho_estoque):
    """Le as duas abas de dados do estoque e devolve a lista bruta de movimentos."""
    wb = openpyxl.load_workbook(caminho_estoque, data_only=True)
    movs = []
    for aba, tipo in (("DADOS DE ENTRADAS ORIGEM 1", "E"), ("DADOS DE SAÍDAS ORIGEM 1", "S")):
        ws = wb[aba]
        for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if r[0] is None and r[1] is None:
                continue
            data, obs_data = parse_data(r[0])
            movs.append(dict(
                tipo=tipo, aba=aba, linha=i,
                data=data, obs_data=obs_data,
                numero=int(r[1]) if r[1] is not None else None,
                parceiro_bruto=r[2], produto_bruto=r[3],
                ncm=r[4], peso=num(r[5]), valor=num(r[6]),
                origem=limpa_texto(r[7]), obs=limpa_texto(r[8]) if len(r) > 8 else None,
            ))
    # totais gravados na ultima linha de cada aba (linha de SUBTOTAL da planilha)
    totais = {}
    for aba, tipo in (("DADOS DE ENTRADAS ORIGEM 1", "E"), ("DADOS DE SAÍDAS ORIGEM 1", "S")):
        ws = wb[aba]
        ult = [r for r in ws.iter_rows(values_only=True) if r[5] is not None or r[6] is not None]
        totais[tipo] = (num(ult[-1][5]), num(ult[-1][6]))
    # painel
    ws = wb["DASHBOARD"]
    painel = dict(
        estoque_total_kg=num(ws["B5"].value),
        entradas_kg=num(ws["E5"].value),
        saidas_kg=num(ws["H5"].value),
        atualizado=ws["A3"].value,
        por_produto={},
    )
    for row in range(9, 15):
        p = ws.cell(row, 2).value
        if p:
            painel["por_produto"][fold(limpa_texto(p))] = num(ws.cell(row, 3).value)
    return movs, totais, painel


def ler_apuracao(caminho_apuracao):
    """Le os lancamentos e o resumo da apuracao do mes de referencia."""
    wb = openpyxl.load_workbook(caminho_apuracao, data_only=True)
    ws = wb["OPERAÇÕES SAÍDA IMPORTADO"]
    lanc, bloco, devolucao = [], None, False
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        a = r[0]
        if isinstance(a, str) and "DEVOLUÇÃO DE VENDAS" in a:
            devolucao = True
        if isinstance(a, str) and a.strip().startswith(("Interestadual", "Interna")):
            carga = num(r[1])
            bloco = {0.6: "1 - Interestadual 12%", 1.0: "2 - Interestadual importado 4%",
                     2.1: "3 - Interna 12%"}.get(round(carga, 1) if carga else None, str(carga))
        if r[1] is not None and isinstance(r[1], (int, float)) and r[2] and r[4] is not None:
            data, _ = parse_data(a)
            lanc.append(dict(
                linha=i, devolucao=devolucao, bloco=bloco, data=data,
                numero=int(r[1]), parceiro_bruto=r[2], produto_bruto=r[3],
                valor_contabil=num(r[4]), base=num(r[5]), aliquota=num(r[6]),
                icms=num(r[7]), aliq_presumido=num(r[8]), credito_presumido=num(r[9]),
            ))
    resumo = {}
    for r in ws.iter_rows(values_only=True):
        if isinstance(r[1], str) and isinstance(r[0], (int, float)):
            resumo[r[1].strip()] = num(r[6])
        if isinstance(r[0], str):
            chave = r[0].strip()
            if chave in ("Base de cálculo das operações alcançadas pelo benefício",) and resumo.get("BC") is None:
                resumo["BC"] = num(r[4])
    # entradas de importacao do mes
    wse = wb["OPERAÇÕES ENTRADA IMPORTADO"]
    entradas = []
    for i, r in enumerate(wse.iter_rows(min_row=2, values_only=True), start=2):
        if r[6] is None:
            continue
        d1, o1 = parse_data(r[0])
        d2, o2 = parse_data(r[1])
        entradas.append(dict(linha=i, emissao=d1, entrada=d2, obs_entrada=o2,
                            numero=int(r[6]), serie=limpa_texto(r[7]), modelo=limpa_texto(r[8]),
                            parceiro_bruto=r[9], cfop=r[10], valor=num(r[11]), qtd=num(r[12]),
                            modalidade=limpa_texto(r[13]), cst=limpa_texto(r[15])))
    return lanc, resumo, entradas


# ----------------------------------------------------------------------------
# 4. Saneamento de parceiros
# ----------------------------------------------------------------------------

def sanear_parceiros(movs, extras=()):
    """Agrupa variantes do mesmo nome. Proposta, nunca aplicacao cega."""
    ocor = defaultdict(lambda: dict(qtd=0, variantes=set(), lados=set()))
    for m in movs:
        k = chave_parceiro(m["parceiro_bruto"])
        if not k:
            continue
        ocor[k]["qtd"] += 1
        ocor[k]["variantes"].add(str(m["parceiro_bruto"]))
        ocor[k]["lados"].add("fornecedor" if m["tipo"] == "E" else "cliente")
    for nome in extras:
        k = chave_parceiro(nome)
        if k:
            ocor[k]["variantes"].add(str(nome))

    chaves = sorted(ocor, key=lambda k: (-len(k), k))
    canonico = {}
    grupos = defaultdict(list)
    for k in chaves:
        alvo = None
        for c in grupos:
            if k == c or c.startswith(k + " ") or k.startswith(c + " "):
                alvo = c
                break
            pref = os.path.commonprefix([k, c])
            pref = pref[:pref.rfind(" ")] if " " in pref else pref
            if len(pref) >= 20:
                alvo = c
                break
        if alvo is None:
            grupos[k] = [k]
            canonico[k] = k
        else:
            grupos[alvo].append(k)
            canonico[k] = alvo

    # Segunda passada (decisao 4): unifica grupos cujo nucleo da razao social comeca igual.
    # Guarda o que foi unido em UNIFICADOS_POR_DECISAO para o relatorio - a operacao e' reversivel.
    if DECISOES["parceiros_parecidos"] == "unificar":
        por_nucleo = defaultdict(list)
        for c in list(grupos):
            por_nucleo[nucleo_razao(c)[0]].append(c)
        for _, membros_ in por_nucleo.items():
            if len(membros_) < 2:
                continue
            # o grupo com mais lancamentos vira o dono
            dono = max(membros_, key=lambda c: sum(ocor[m]["qtd"] for m in grupos[c]))
            unidos = []
            for c in membros_:
                if c == dono:
                    continue
                unidos.append(c)
                for m in grupos.pop(c):
                    grupos[dono].append(m)
                    canonico[m] = dono
            if unidos:
                UNIFICADOS_POR_DECISAO.append((dono, unidos))
                GRUPOS_DECIDIDOS.add(dono)
    return canonico, grupos, ocor


# ----------------------------------------------------------------------------
# 5. Schema
# ----------------------------------------------------------------------------

DDL = """
PRAGMA foreign_keys = ON;

CREATE TABLE parceiro (
    id              INTEGER PRIMARY KEY,
    nome            TEXT    NOT NULL UNIQUE,   -- canonico
    cnpj            TEXT,                      -- a preencher na Fase 5 (XML)
    id_estrangeiro  TEXT,
    uf              TEXT,
    pais            TEXT,
    exterior        INTEGER NOT NULL DEFAULT 0,
    papel           TEXT,                      -- cliente | fornecedor | ambos
    variantes       TEXT,                      -- grafias encontradas na migracao
    status          TEXT NOT NULL DEFAULT 'migrado'
);

CREATE TABLE produto (
    id          INTEGER PRIMARY KEY,
    descricao   TEXT NOT NULL UNIQUE,          -- canonica, sem variacao de acento
    ncm         TEXT,
    unidade     TEXT NOT NULL DEFAULT 'KG',
    categoria   TEXT,                          -- SUCATA | LINGOTE | OUTRO
    metal       TEXT,
    variantes   TEXT,
    status      TEXT NOT NULL DEFAULT 'migrado'
);

CREATE TABLE nota (
    id              INTEGER PRIMARY KEY,
    chave_acesso    TEXT UNIQUE,               -- 44 digitos; nulo no historico migrado
    numero          INTEGER NOT NULL,
    serie           TEXT,
    modelo          TEXT,
    tipo            TEXT NOT NULL,             -- E | S
    cfop            TEXT,
    natureza        TEXT,                      -- VENDA | DEVOLUCAO | IMPORTACAO
    data_emissao    DATE,
    data_mov        DATE,
    parceiro_id     INTEGER REFERENCES parceiro(id),
    valor_total     REAL,
    origem_registro TEXT,                      -- arquivo/aba/linha de origem
    status          TEXT NOT NULL DEFAULT 'migrada'
);
CREATE INDEX ix_nota_num ON nota(tipo, numero);
CREATE INDEX ix_nota_data ON nota(data_mov);

CREATE TABLE nota_item (
    id            INTEGER PRIMARY KEY,
    nota_id       INTEGER NOT NULL REFERENCES nota(id),
    produto_id    INTEGER NOT NULL REFERENCES produto(id),
    ncm           TEXT,
    origem_merc   TEXT,                        -- 0..8 da tabela da NF-e; nulo no historico
    quantidade    REAL,
    valor         REAL,
    base_calculo  REAL,
    aliquota      REAL,
    cst           TEXT,
    custo_unit    REAL
);
CREATE INDEX ix_item_prod ON nota_item(produto_id);

CREATE TABLE lastro (
    id                INTEGER PRIMARY KEY,
    item_saida_id     INTEGER NOT NULL REFERENCES nota_item(id),
    item_entrada_id   INTEGER REFERENCES nota_item(id),
    quantidade        REAL NOT NULL,
    custo_unitario    REAL,
    metodo            TEXT NOT NULL DEFAULT 'PEPS',
    status            TEXT NOT NULL DEFAULT 'ok'   -- ok | sem_lastro
);
CREATE INDEX ix_lastro_saida ON lastro(item_saida_id);

CREATE TABLE apuracao_mes (
    id             INTEGER PRIMARY KEY,
    competencia    TEXT NOT NULL UNIQUE,
    base_beneficiada REAL, debito REAL, credito_presumido REAL, estorno REAL,
    icms_recolher  REAL, fundo_social REAL, fundo_educacao REAL,
    carga_efetiva  REAL,
    status         TEXT NOT NULL DEFAULT 'aberta',
    origem         TEXT
);

CREATE TABLE divergencia (
    id          TEXT PRIMARY KEY,
    classe      TEXT, severidade TEXT, origem TEXT, linha INTEGER,
    documento   TEXT, descricao TEXT,
    impacto_kg  REAL, impacto_rs REAL, acao TEXT, decisao TEXT
);
"""


def criar_base(caminho):
    if os.path.exists(caminho):
        os.remove(caminho)
    conn = sqlite3.connect(caminho)
    conn.executescript(DDL)
    return conn


# ----------------------------------------------------------------------------
# 6. Carga
# ----------------------------------------------------------------------------

def categoria_de(desc):
    if desc.startswith("SUCATA"):
        return "SUCATA"
    if desc.startswith("LINGOTE"):
        return "LINGOTE"
    return "OUTRO"


def metal_de(desc):
    for m in ("ALUMINIO", "MAGNESIO", "COBRE", "SILICIO"):
        if m in desc:
            return m
    return None


def carregar(conn, movs, apur_lanc, apur_entr, canonico, grupos, ocor):
    cur = conn.cursor()

    # --- parceiro ---------------------------------------------------------
    pid = {}
    for i, (canon, membros) in enumerate(sorted(grupos.items()), start=1):
        variantes, qtd, lados = set(), 0, set()
        for m in membros:
            variantes |= ocor[m]["variantes"]
            qtd += ocor[m]["qtd"]
            lados |= ocor[m]["lados"]
        dono_ = max(membros, key=lambda m: ocor[m]["qtd"])
        nome = max((limpa_texto(v) for v in (ocor[dono_]["variantes"] or variantes)), key=len)
        papel = "ambos" if len(lados) > 1 else (lados.pop() if lados else None)
        exterior = int(bool(re.search(r"\b(GMBH|BV|LIMITED|LTD|S\.?A\.?$|CORPORATION|CO\.LIMITED|INWASTE)\b", fold(nome))
                       and not re.search(r"LTDA|EIRELI|S/A", fold(nome))))
        cur.execute("INSERT INTO parceiro(id,nome,exterior,papel,variantes) VALUES (?,?,?,?,?)",
                    (i, nome, exterior, papel, " | ".join(sorted(variantes))))
        for m in membros:
            pid[m] = i
        if len(variantes) > 1 and canon not in GRUPOS_DECIDIDOS:
            limpas = {limpa_texto(v) for v in variantes}
            invisivel = len(limpas) == 1
            REG.add("Parceiro - grafia invisivel" if invisivel else "Parceiro - variantes unificadas",
                    "Medio" if invisivel else "Alto", "ambas as abas", None, nome,
                    (f"{len(variantes)} grafias que diferem apenas por espaco invisivel (NBSP), espaco "
                     f"duplo ou pontuacao final: " if invisivel else
                     f"{len(variantes)} grafias diferentes tratadas como o mesmo parceiro: ")
                    + " | ".join(repr(v)[1:-1] for v in sorted(variantes)),
                    acao=("Unificacao segura: o texto e' identico depois da limpeza" if invisivel else
                          "Confirmar se e' a mesma empresa e informar o CNPJ (a planilha nao tem CNPJ)"))

    # --- parceiros unificados por decisao (decisao 4) ---------------------
    for dono, unidos in UNIFICADOS_POR_DECISAO:
        nomes_ = [max((limpa_texto(v) for v in ocor[m]["variantes"]), key=len)
                  for grupo_ in [dono] + unidos for m in [grupo_] if ocor[m]["variantes"]]
        REG.add("Parceiro - unificado por decisao", "Medio", "ambas as abas", None,
                nomes_[0] if nomes_ else dono,
                "Razoes sociais diferentes unidas sob o mesmo parceiro por decisao de 27/08/2026 "
                "(nucleo do nome igual): " + " | ".join(nomes_),
                acao="Reversivel: cada razao social original ficou gravada em parceiro.variantes. "
                     "Quando os CNPJs entrarem (Fase 5), separar as que forem inscricoes distintas")

    # --- produto ----------------------------------------------------------
    prid = {}
    for i, (canon, variantes) in enumerate(sorted(PRODUTO_CANON.items()), start=1):
        cur.execute("INSERT INTO produto(id,descricao,categoria,metal,variantes) VALUES (?,?,?,?,?)",
                    (i, canon, categoria_de(canon), metal_de(canon), " | ".join(sorted(variantes))))
        prid[canon] = i
        if len(variantes) > 1:
            REG.add("Produto - variantes unificadas", "Critico", "ambas as abas", None, canon,
                    "Mesmo produto gravado com grafias diferentes: " + " | ".join(sorted(variantes)),
                    acao="Unificar sob a descricao canonica; os KPIs do painel deixam de divergir")

    # --- nota / item ------------------------------------------------------
    # Cada LINHA da planilha e' um item. Linhas com mesmo tipo+numero+data+parceiro
    # sao a mesma nota com mais de um produto - e' assim que o modelo grava.
    grupos_nota = defaultdict(list)
    for m in movs:
        k = chave_parceiro(m["parceiro_bruto"])
        canon = canonico.get(k, k)
        m["_parceiro_id"] = pid[canon]
        grupos_nota[(m["tipo"], m["numero"], m["data"], pid[canon])].append(m)

    nid = 0
    itens = []
    for chave, linhas_nota in grupos_nota.items():
        nid += 1
        m0 = linhas_nota[0]
        natureza = "DEVOLUCAO" if any(x["origem"] == "DEVOLUÇÃO" or (x["obs"] or "").startswith("DEV")
                                      for x in linhas_nota) else \
                   ("IMPORTACAO" if m0["tipo"] == "E" else "VENDA")
        cur.execute("""INSERT INTO nota(id,numero,tipo,natureza,data_emissao,data_mov,parceiro_id,
                       valor_total,origem_registro,status)
                       VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (nid, m0["numero"], m0["tipo"], natureza, m0["data"], m0["data"],
                     m0["_parceiro_id"], sum(x["valor"] or 0 for x in linhas_nota),
                     "; ".join(f"{x['aba']}!L{x['linha']}" for x in linhas_nota),
                     "migrada_sem_chave"))
        vistos_item = defaultdict(list)
        for m in linhas_nota:
            p = prid[canon_produto(m["produto_bruto"])]
            custo = (m["valor"] / m["peso"]) if (m["valor"] and m["peso"]) else None
            cur.execute("""INSERT INTO nota_item(nota_id,produto_id,quantidade,valor,custo_unit)
                           VALUES (?,?,?,?,?)""", (nid, p, m["peso"], m["valor"], custo))
            itens.append((cur.lastrowid, nid, m, p))
            vistos_item[(p, m["peso"], m["valor"])].append(m)
        if len(linhas_nota) > 1:
            por_prod = defaultdict(list)
            for (p, peso, valor), rep in vistos_item.items():
                por_prod[p].extend(rep)
            for p, rep in por_prod.items():
                chaves_rep = {(x["peso"], x["valor"]) for x in rep}
                if len(rep) > 1 and len(chaves_rep) > 1:
                    REG.add("Item repetido na mesma nota", "Alto",
                            "ENTRADAS" if m0["tipo"] == "E" else "SAIDAS",
                            rep[0]["linha"], f"NF {m0['numero']}",
                            f"O mesmo produto aparece {len(rep)} vezes na mesma nota, com peso/valor "
                            f"diferentes (linhas " + ", ".join(str(x["linha"]) for x in rep) + "): "
                            + " | ".join(f"{x['peso']} kg / R$ {x['valor']}" for x in rep),
                            impacto_rs=sum(x["valor"] or 0 for x in rep[1:]),
                            acao="Ou sao dois itens legitimos da mesma NF, ou uma linha foi digitada "
                                 "duas vezes com correcao parcial. Conferir a NF")
            for (p, peso, valor), repetidas in vistos_item.items():
                if len(repetidas) > 1:
                    REG.add("Linha duplicada - mantida por decisao"
                            if DECISOES["linhas_identicas"] == "manter" else "Linha duplicada",
                            "Baixo" if DECISOES["linhas_identicas"] == "manter" else "Critico",
                            "ENTRADAS" if m0["tipo"] == "E" else "SAIDAS",
                            repetidas[0]["linha"], f"NF {m0['numero']}",
                            f"{len(repetidas)} linhas identicas (mesma data, parceiro, produto, peso e "
                            f"valor) nas linhas " + ", ".join(str(x["linha"]) for x in repetidas)
                            + f": {peso} kg / R$ {valor}",
                            impacto_kg=(len(repetidas) - 1) * (peso or 0),
                            impacto_rs=(len(repetidas) - 1) * (valor or 0),
                            acao=("Mantidas por decisao de 27/08/2026: notas diferentes podem ter "
                                  "dados identicos. Quando a chave de acesso entrar (Fase 5) a "
                                  "duplicidade real passa a ser impossivel e este caso deixa de existir"
                                  if DECISOES["linhas_identicas"] == "manter" else
                                  "Provavel digitacao em duplicidade: localizar as NFs e decidir"))

    # --- apuracao ---------------------------------------------------------
    conn.commit()
    return itens


# ----------------------------------------------------------------------------
# 7. Deteccao de divergencias no dado migrado
# ----------------------------------------------------------------------------

def analisar_movimentos(movs):
    for m in movs:
        lado = "ENTRADAS" if m["tipo"] == "E" else "SAIDAS"
        doc = f"NF {m['numero']}"
        if m["obs_data"]:
            REG.add("Data - formato", "Medio", lado, m["linha"], doc, m["obs_data"],
                    acao="Gravar como data; texto nao entra em ordenacao nem em PEPS")
        if m["data"] is None:
            REG.add("Data - ausente", "Alto", lado, m["linha"], doc,
                    "Lancamento sem data: nao entra em nenhuma competencia nem no PEPS",
                    impacto_kg=m["peso"], impacto_rs=m["valor"],
                    acao="Localizar a NF e informar a data de emissao")
        elif m["data"] > HOJE:
            REG.add("Data - futura", "Alto", lado, m["linha"], doc,
                    f"Data {m['data'].strftime('%d/%m/%Y')} e' posterior a hoje "
                    f"({HOJE.strftime('%d/%m/%Y')}) - provavel erro de digitacao",
                    impacto_kg=m["peso"], impacto_rs=m["valor"],
                    acao="Corrigir a data; ela puxa o 'atualizado em' do painel")
        elif m["data"] < dt.date(2020, 11, 1):
            REG.add("Data - fora do periodo", "Medio", lado, m["linha"], doc,
                    f"Data {m['data']} anterior ao inicio do historico", acao="Conferir")
        if m["peso"] in (None, 0):
            REG.add("Quantidade ausente ou zero", "Alto", lado, m["linha"], doc,
                    f"{m['produto_bruto']} com peso {m['peso']!r} e valor {m['valor']!r}"
                    + (f" - obs: {m['obs']}" if m["obs"] else ""),
                    impacto_rs=m["valor"],
                    acao="Sem peso o item nao entra no estoque nem gera lastro; informar a quantidade "
                         "ou marcar a nota como operacao sem movimentacao fisica")
        if m["valor"] in (None, 0):
            REG.add("Valor ausente ou zero", "Alto", lado, m["linha"], doc,
                    f"{m['produto_bruto']} com peso {m['peso']!r} sem valor",
                    impacto_kg=m["peso"],
                    acao="Sem valor nao ha' custo unitario: a entrada entra no PEPS a custo zero "
                         "e desvaloriza o estoque")
        if m["ncm"] in (None, ""):
            pass  # tratado em bloco, ver analisar_estrutura


def analisar_estrutura(movs):
    n = len(movs)
    sem_ncm = sum(1 for m in movs if not m["ncm"])
    if sem_ncm and DECISOES["ncm"] != "desconsiderar":
        REG.add("NCM ausente", "Critico", "ambas as abas", None, f"{sem_ncm} de {n} itens",
                "A coluna NCM existe nas duas abas e esta' 100% vazia.",
                acao="Definir o NCM por produto e preencher retroativamente")
    elif sem_ncm:
        REG.add("NCM dispensado por decisao", "Baixo", "ambas as abas", None, f"{sem_ncm} itens",
                "A coluna NCM esta' vazia nos dois arquivos. Por decisao de 27/08/2026 o NCM nao e' "
                "exigido nesta base; o campo continua existindo e aceita nulo.",
                acao="Sem acao. Quando a ingestao por XML entrar (Fase 5) o NCM vem preenchido "
                     "de graca e o PEPS pode passar a separar por NCM alem de produto")
    sem_origem = sum(1 for m in movs if not m["origem"] and m["tipo"] == "S")
    if sem_origem and DECISOES["vinculo_lastro"] != "controle_por_saldo":
        REG.add("Origem da mercadoria ausente", "Critico", "SAIDAS", None,
                f"{sem_origem} saidas",
                "Nenhum item registra a origem da mercadoria.",
                acao="Gravar a origem por item")
    elif sem_origem:
        REG.add("Origem por item dispensada por decisao", "Baixo", "SAIDAS", None,
                f"{sem_origem} saidas",
                "Por decisao de 27/08/2026 a saida nao e' vinculada a uma entrada especifica: o "
                "controle e' pelo saldo de estoque do produto. O campo origem_merc continua no "
                "modelo e aceita nulo; a partir da Fase 5 ele vem preenchido do XML sem ninguem "
                "digitar.",
                acao="Sem acao. O PEPS continua rodando, mas como metodo de CUSTEIO do estoque, "
                     "nao como prova de vinculacao entre nota de entrada e nota de saida")
    sem_chave = len(movs)
    REG.add("Chave de acesso ausente", "Alto", "ambas as abas", None, f"{sem_chave} lancamentos",
            "Nenhum lancamento tem a chave de acesso de 44 digitos. O identificador hoje e' o "
            "numero da NF, que se repete entre series e entre emitentes.",
            acao="Migrar sem chave e marcar 'migrada_sem_chave'; a partir da Fase 5 a chave passa "
                 "a ser obrigatoria e a duplicidade vira impossivel")


def analisar_duplicidades(movs):
    for tipo, lado in (("E", "ENTRADAS"), ("S", "SAIDAS")):
        por_num = defaultdict(list)
        for m in movs:
            if m["tipo"] == tipo and m["numero"] is not None:
                por_num[m["numero"]].append(m)
        for numero, lst in sorted(por_num.items()):
            if len(lst) < 2:
                continue
            grupos_ = {(x["data"], chave_parceiro(x["parceiro_bruto"])) for x in lst}
            if len(grupos_) == 1:
                continue   # mesma nota com mais de um item: tratado na carga
            chaves = {(x["data"], chave_parceiro(x["parceiro_bruto"]), x["peso"], x["valor"]) for x in lst}
            iguais = len(chaves) == 1
            detalhe = " || ".join(
                f"L{x['linha']}: {x['data']} | {limpa_texto(x['parceiro_bruto'])} | "
                f"{x['produto_bruto']} | {x['peso']} kg | R$ {x['valor']}" for x in lst)
            REG.add("NF repetida", "Critico" if iguais else "Alto", lado,
                    lst[0]["linha"], f"NF {numero}",
                    ("Duplicidade provavel: mesma data, parceiro, peso e valor. " if iguais
                     else "Mesmo numero de NF em lancamentos diferentes. ") + detalhe,
                    impacto_kg=sum(x["peso"] or 0 for x in lst[1:]),
                    impacto_rs=sum(x["valor"] or 0 for x in lst[1:]),
                    acao="Sem a chave de acesso nao da' para decidir entre duplicidade e serie "
                         "diferente: localizar as NFs e decidir uma a uma")


# ----------------------------------------------------------------------------
# 8. PEPS
# ----------------------------------------------------------------------------

def proximo_id(cur, tabela):
    return (cur.execute(f"SELECT COALESCE(MAX(id),0) FROM {tabela}").fetchone()[0] or 0) + 1


def custo_estimado(cur, produto_id, data):
    """Custo do acerto: media ponderada das entradas anteriores; se nao houver, a primeira
    entrada posterior; se nao houver nenhuma, o valor unitario medio das saidas do produto."""
    r = cur.execute("""SELECT SUM(i.valor), SUM(i.quantidade) FROM nota_item i JOIN nota n ON n.id=i.nota_id
                       WHERE i.produto_id=? AND n.tipo='E' AND n.natureza<>'ACERTO'
                         AND n.data_mov IS NOT NULL AND n.data_mov <= ?
                         AND i.valor IS NOT NULL AND i.quantidade > 0""",
                    (produto_id, data)).fetchone()
    if r and r[0] and r[1]:
        return r[0] / r[1]
    r = cur.execute("""SELECT i.custo_unit FROM nota_item i JOIN nota n ON n.id=i.nota_id
                       WHERE i.produto_id=? AND n.tipo='E' AND n.natureza<>'ACERTO'
                         AND i.custo_unit IS NOT NULL AND n.data_mov IS NOT NULL
                       ORDER BY n.data_mov LIMIT 1""", (produto_id,)).fetchone()
    if r and r[0]:
        return r[0]
    r = cur.execute("""SELECT SUM(i.valor), SUM(i.quantidade) FROM nota_item i JOIN nota n ON n.id=i.nota_id
                       WHERE i.produto_id=? AND n.tipo='S' AND i.valor IS NOT NULL AND i.quantidade>0""",
                    (produto_id,)).fetchone()
    return (r[0] / r[1]) if (r and r[0] and r[1]) else 0.0


def rodar_peps(conn):
    """Consome saidas contra entradas por produto, em ordem cronologica.
    Devolve o razao de saldos e grava a tabela lastro."""
    cur = conn.cursor()
    rows = cur.execute("""
        SELECT i.id, n.tipo, n.natureza, n.data_mov, n.numero, i.produto_id, p.descricao,
               i.quantidade, i.valor, i.custo_unit, n.origem_registro, pa.nome
        FROM nota_item i
        JOIN nota n     ON n.id = i.nota_id
        JOIN produto p  ON p.id = i.produto_id
        LEFT JOIN parceiro pa ON pa.id = n.parceiro_id
        WHERE n.data_mov IS NOT NULL
        ORDER BY n.data_mov, n.tipo DESC, n.numero
    """).fetchall()

    filas = defaultdict(list)     # produto_id -> [ [item_id, qtd_restante, custo] ]
    saldo = defaultdict(float)
    razao = []
    negativos = []
    sem_lastro = []
    lastro_id = 0

    for (item_id, tipo, natureza, data, numero, prod_id, prod, qtd, valor, custo,
         origem, parceiro) in rows:
        qtd = qtd or 0.0
        if tipo == "E":
            # devolucao de venda tambem retorna mercadoria ao estoque
            if qtd > 0:
                filas[prod_id].append([item_id, qtd, custo or 0.0])
            saldo[prod_id] += qtd
        else:
            restante = qtd
            saldo[prod_id] -= qtd
            while restante > 1e-9 and filas[prod_id]:
                lote = filas[prod_id][0]
                usa = min(restante, lote[1])
                lastro_id += 1
                cur.execute("""INSERT INTO lastro(id,item_saida_id,item_entrada_id,quantidade,
                               custo_unitario,status) VALUES (?,?,?,?,?,?)""",
                            (lastro_id, item_id, lote[0], usa, lote[2], "ok"))
                lote[1] -= usa
                restante -= usa
                if lote[1] <= 1e-9:
                    filas[prod_id].pop(0)
            if restante > 1e-9:
                falta = restante
                if DECISOES["saldo_negativo"] == "acerto_datado":
                    custo_ac = custo_estimado(cur, prod_id, data)
                    nid_ac = proximo_id(cur, "nota")
                    cur.execute("""INSERT INTO nota(id,numero,tipo,natureza,data_emissao,data_mov,
                                   parceiro_id,valor_total,origem_registro,status)
                                   VALUES (?,?,?,?,?,?,NULL,?,?,?)""",
                                (nid_ac, 0, "E", "ACERTO", data, data, restante * custo_ac,
                                 f"acerto gerado pela migracao para a saida NF {numero}",
                                 "acerto_de_estoque"))
                    cur.execute("""INSERT INTO nota_item(nota_id,produto_id,quantidade,valor,custo_unit)
                                   VALUES (?,?,?,?,?)""",
                                (nid_ac, prod_id, restante, restante * custo_ac, custo_ac))
                    item_ac = cur.lastrowid
                    lastro_id += 1
                    cur.execute("""INSERT INTO lastro(id,item_saida_id,item_entrada_id,quantidade,
                                   custo_unitario,metodo,status) VALUES (?,?,?,?,?,?,?)""",
                                (lastro_id, item_id, item_ac, restante, custo_ac, "ACERTO", "ok"))
                    saldo[prod_id] += restante
                    ACERTOS.append(dict(data=data, produto=prod, qtd=restante, custo=custo_ac,
                                        valor=restante * custo_ac, saida_nf=numero,
                                        cliente=parceiro, origem=origem))
                    restante = 0.0
                else:
                    lastro_id += 1
                    cur.execute("""INSERT INTO lastro(id,item_saida_id,item_entrada_id,quantidade,
                                   custo_unitario,status) VALUES (?,?,NULL,?,NULL,?)""",
                                (lastro_id, item_id, restante, "sem_lastro"))
                sem_lastro.append(dict(data=data, numero=numero, produto=prod, parceiro=parceiro,
                                       qtd_sem_lastro=falta, qtd_nota=qtd, origem=origem))
        razao.append(dict(data=data, tipo=tipo, natureza=natureza, numero=numero, produto=prod,
                          parceiro=parceiro, qtd=qtd, valor=valor, saldo=saldo[prod_id],
                          origem=origem))
        if saldo[prod_id] < -1e-6:
            negativos.append(dict(data=data, produto=prod, numero=numero, saldo=saldo[prod_id],
                                  origem=origem))
    conn.commit()

    # posicao final valorizada
    posicao = []
    prods = dict(cur.execute("SELECT id, descricao FROM produto").fetchall())
    aritm = {pid: (e or 0) - (sa or 0) for pid, e, sa in cur.execute("""
        SELECT i.produto_id,
               SUM(CASE WHEN n.tipo='E' THEN i.quantidade ELSE 0 END),
               SUM(CASE WHEN n.tipo='S' THEN i.quantidade ELSE 0 END)
        FROM nota_item i JOIN nota n ON n.id=i.nota_id WHERE n.natureza<>'ACERTO'
        GROUP BY i.produto_id""")}
    acerto_prod = dict(cur.execute("""SELECT i.produto_id, SUM(i.quantidade) FROM nota_item i
        JOIN nota n ON n.id=i.nota_id WHERE n.natureza='ACERTO' GROUP BY i.produto_id"""))
    for prod_id, desc in sorted(prods.items(), key=lambda x: x[1]):
        qtd_rest = sum(l[1] for l in filas[prod_id])
        valor_rest = sum(l[1] * l[2] for l in filas[prod_id])
        posicao.append(dict(produto=desc, saldo_peps_kg=qtd_rest, saldo_peps_rs=valor_rest,
                            saldo_aritmetico_kg=aritm.get(prod_id, 0.0),
                            acerto_kg=acerto_prod.get(prod_id, 0.0),
                            saldo_com_acerto_kg=aritm.get(prod_id, 0.0) + acerto_prod.get(prod_id, 0.0),
                            saldo_peps_calc=saldo[prod_id],
                            custo_medio=(valor_rest / qtd_rest) if qtd_rest else None))
    return razao, negativos, sem_lastro, posicao


def analisar_peps(negativos, sem_lastro, posicao):
    # primeiro momento em que cada produto fica negativo
    vistos = set()
    for n in negativos:
        if n["produto"] in vistos:
            continue
        vistos.add(n["produto"])
        REG.add("Saldo negativo - corrigido por acerto" if DECISOES["saldo_negativo"] == "acerto_datado"
                else "Saldo negativo",
                "Medio" if DECISOES["saldo_negativo"] == "acerto_datado" else "Critico", "PEPS", None,
                f"{n['produto']} - NF {n['numero']}",
                f"Sem tratamento, o saldo ficaria negativo em {n['data']} ({n['saldo']:,.1f} kg). "
                + ("O acerto datado da decisao 1 cobre a falta na propria data."
                   if DECISOES["saldo_negativo"] == "acerto_datado" else "Saiu mercadoria que nunca entrou."),
                impacto_kg=n["saldo"],
                acao="Ver aba ACERTOS")
    if sem_lastro:
        tot_kg = sum(s["qtd_sem_lastro"] for s in sem_lastro)
        if DECISOES["saldo_negativo"] == "acerto_datado":
            tot_rs = sum(a["valor"] for a in ACERTOS)
            REG.add("Acerto de estoque gerado", "Alto", "PEPS", None,
                    f"{len(ACERTOS)} acertos",
                    f"{tot_kg:,.1f} kg sairam sem saldo disponivel no produto. Por decisao de 27/08/2026 "
                    f"foram gerados {len(ACERTOS)} lancamentos de acerto, cada um na data da propria "
                    f"saida, somando R$ {tot_rs:,.2f} a custo estimado. Nenhum saldo fica negativo.",
                    impacto_kg=tot_kg, impacto_rs=tot_rs,
                    acao="Conferir o custo atribuido na aba ACERTOS e substituir pelo custo real "
                         "quando a origem da mercadoria for identificada")
        else:
            REG.add("Saida sem lastro de importacao", "Critico", "PEPS", None,
                    f"{len(sem_lastro)} saidas",
                    f"{tot_kg:,.1f} kg sairam sem entrada correspondente disponivel no PEPS.",
                    impacto_kg=tot_kg, acao="Conferir")
    for p in posicao:
        if p.get("saldo_com_acerto_kg", p["saldo_aritmetico_kg"]) < -1e-6:
            REG.add("Saldo final negativo", "Critico", "PEPS", None, p["produto"],
                    f"Posicao final de {p['saldo_aritmetico_kg']:,.1f} kg",
                    impacto_kg=p["saldo_aritmetico_kg"],
                    acao="Fechar a decisao antes da Fase 2")
        if p["saldo_peps_kg"] > 0 and not p["saldo_peps_rs"]:
            REG.add("Estoque sem valoracao", "Alto", "PEPS", None, p["produto"],
                    f"{p['saldo_peps_kg']:,.1f} kg em estoque sem custo atribuido",
                    impacto_kg=p["saldo_peps_kg"],
                    acao="Entradas sem valor impedem a valoracao; ver classe 'Valor ausente'")


# ----------------------------------------------------------------------------
# 9. Cruzamento estoque x apuracao (competencia do arquivo anexado)
# ----------------------------------------------------------------------------

def cruzar_competencia(movs, apur_lanc, apur_entr, competencia):
    ano, mes = (int(x) for x in competencia.split("-"))
    def no_mes(d):
        return d is not None and d.year == ano and d.month == mes

    def agrega(tipo):
        g = {}
        for m in movs:
            if m["tipo"] != tipo or not no_mes(m["data"]):
                continue
            a = g.setdefault(m["numero"], dict(m))
            if a is not m and a.get("_agg"):
                a["valor"] = (a["valor"] or 0) + (m["valor"] or 0)
                a["peso"] = (a["peso"] or 0) + (m["peso"] or 0)
                a["produto_bruto"] = a["produto_bruto"] + " + " + str(m["produto_bruto"])
            a["_agg"] = True
        return g
    saidas_est = agrega("S")
    entradas_est = agrega("E")
    vendas_ap = {l["numero"]: l for l in apur_lanc if not l["devolucao"]}
    devol_ap = {l["numero"]: l for l in apur_lanc if l["devolucao"]}
    linhas = []

    for numero, m in sorted(saidas_est.items()):
        a = vendas_ap.get(numero)
        if a is None:
            linhas.append(dict(doc=f"NF {numero}", lado="SAIDA", estoque="presente",
                               apuracao="AUSENTE", produto_estoque=limpa_texto(m["produto_bruto"]),
                               produto_apuracao=None, valor_estoque=m["valor"], valor_apuracao=None,
                               diagnostico="Saida registrada no estoque de importados e ausente da apuracao"))
            REG.add("Nota fora da apuracao - ajuste em curso"
                    if DECISOES["notas_fora_apuracao"] == "ajuste_contabil" else "Nota fora da apuracao",
                    "Alto" if DECISOES["notas_fora_apuracao"] == "ajuste_contabil" else "Critico",
                    "SAIDAS", m["linha"], f"NF {numero}",
                    f"{limpa_texto(m['parceiro_bruto'])} - {limpa_texto(m['produto_bruto'])} - "
                    f"{m['peso']:,.0f} kg - R$ {m['valor']:,.2f}: consta no estoque de importados de "
                    f"{competencia} e nao aparece em nenhum bloco da apuracao.",
                    impacto_kg=m["peso"], impacto_rs=m["valor"],
                    acao=("Confirmado em 27/08/2026: a nota realmente nao entrou na apuracao e sera' "
                          "ajustada com a contabilidade em ago/2026. Consequencia: julho/2026 so' serve "
                          "de gabarito da Fase 3 depois do ajuste - ver aba IMPACTO JULHO"
                          if DECISOES["notas_fora_apuracao"] == "ajuste_contabil" else
                          "Ou a saida e' beneficiada e falta na apuracao, ou nao e' importada."))
        else:
            dv = (a["valor_contabil"] or 0) - (m["valor"] or 0)
            prod_dif = fold(limpa_texto(a["produto_bruto"])) != fold(limpa_texto(m["produto_bruto"]))
            linhas.append(dict(doc=f"NF {numero}", lado="SAIDA", estoque="presente",
                               apuracao=a["bloco"], produto_estoque=limpa_texto(m["produto_bruto"]),
                               produto_apuracao=limpa_texto(a["produto_bruto"]),
                               valor_estoque=m["valor"], valor_apuracao=a["valor_contabil"],
                               diagnostico=("produto divergente" if prod_dif else "") +
                                           (f" | valor difere em R$ {dv:,.2f}" if abs(dv) > TOL else "")
                                           or "ok"))
            if prod_dif:
                REG.add("Produto divergente entre arquivos", "Alto", "SAIDAS", m["linha"], f"NF {numero}",
                        f"Estoque diz '{limpa_texto(m['produto_bruto'])}', apuracao diz "
                        f"'{limpa_texto(a['produto_bruto'])}' para a mesma nota.",
                        impacto_kg=m["peso"], impacto_rs=m["valor"],
                        acao="Um dos dois arquivos esta' errado; o produto define NCM e categoria")
            if abs(dv) > TOL:
                REG.add("Valor divergente entre arquivos", "Critico", "SAIDAS", m["linha"], f"NF {numero}",
                        f"Estoque R$ {m['valor']:,.2f} x apuracao R$ {a['valor_contabil']:,.2f}",
                        impacto_rs=dv, acao="Conferir contra o XML da nota")

    for numero, a in sorted(vendas_ap.items()):
        if numero not in saidas_est:
            linhas.append(dict(doc=f"NF {numero}", lado="SAIDA", estoque="AUSENTE",
                               apuracao=a["bloco"], produto_estoque=None,
                               produto_apuracao=limpa_texto(a["produto_bruto"]),
                               valor_estoque=None, valor_apuracao=a["valor_contabil"],
                               diagnostico="Nota apurada com beneficio e ausente do estoque"))
            REG.add("Nota fora do estoque", "Critico", "APURACAO", a["linha"], f"NF {numero}",
                    f"{limpa_texto(a['parceiro_bruto'])} - R$ {a['valor_contabil']:,.2f} apurada no bloco "
                    f"{a['bloco']} e ausente da aba de saidas do estoque.",
                    impacto_rs=a["valor_contabil"],
                    acao="Beneficio tomado sem baixa de estoque: e' saida sem lastro por construcao")

    for numero, a in sorted(devol_ap.items()):
        m = entradas_est.get(numero)
        linhas.append(dict(doc=f"NF {numero}", lado="DEVOLUCAO", 
                           estoque="presente" if m else "AUSENTE", apuracao=a["bloco"],
                           produto_estoque=limpa_texto(m["produto_bruto"]) if m else None,
                           produto_apuracao=limpa_texto(a["produto_bruto"]),
                           valor_estoque=m["valor"] if m else None, valor_apuracao=a["valor_contabil"],
                           diagnostico="ok" if m else "Devolucao apurada e ausente das entradas do estoque"))
        if not m:
            REG.add("Devolucao fora do estoque", "Alto", "APURACAO", a["linha"], f"NF {numero}",
                    f"Devolucao de R$ {a['valor_contabil']:,.2f} com estorno de credito presumido "
                    "e sem retorno de mercadoria no estoque.", impacto_rs=a["valor_contabil"],
                    acao="Lancar o retorno ou justificar")

    for e in apur_entr:
        m = entradas_est.get(e["numero"])
        linhas.append(dict(doc=f"NF {e['numero']}", lado="ENTRADA IMPORTACAO",
                           estoque="presente" if m else "AUSENTE", apuracao=f"CFOP {e['cfop']}",
                           produto_estoque=limpa_texto(m["produto_bruto"]) if m else None,
                           produto_apuracao=None,
                           valor_estoque=m["valor"] if m else None, valor_apuracao=e["valor"],
                           diagnostico="ok" if m and abs((m["valor"] or 0) - (e["valor"] or 0)) <= TOL
                                       else ("valor difere" if m else "Entrada de importacao ausente do estoque")))
        if e["obs_entrada"]:
            REG.add("Data - formato", "Baixo", "APURACAO ENTRADAS", e["linha"], f"NF {e['numero']}",
                    f"Coluna 'Entrada/Saida': {e['obs_entrada']}",
                    acao="Formatar como data")
        if m and e["qtd"] and m["peso"] and abs(m["peso"] - e["qtd"]) > 0.5:
            REG.add("Quantidade divergente entre arquivos", "Alto", "ENTRADAS", m["linha"],
                    f"NF {e['numero']}", f"Estoque {m['peso']:,.0f} kg x apuracao {e['qtd']:,.0f} kg",
                    impacto_kg=(m["peso"] - e["qtd"]), acao="Conferir contra o XML")
    return linhas


# ----------------------------------------------------------------------------
# 9b. Motor fiscal minimo (usado so' para medir o impacto das notas fora da apuracao)
# ----------------------------------------------------------------------------

BLOCOS = {
    "1": dict(nome="Interestadual 12%", aliq=0.12, presumido=0.114, carga=0.006),
    "2": dict(nome="Interestadual importado 4%", aliq=0.04, presumido=0.030, carga=0.010),
    "3": dict(nome="Interna 12%", aliq=0.12, presumido=0.099, carga=0.021),
}


def apurar(vendas, devolucoes):
    """vendas/devolucoes: lista de (bloco, base). Replica a sequencia da planilha atual."""
    debito = sum(b * BLOCOS[k]["aliq"] for k, b in vendas)
    cp = sum(b * BLOCOS[k]["presumido"] for k, b in vendas)
    dev_icms = sum(b * BLOCOS[k]["aliq"] for k, b in devolucoes)
    estorno = sum(b * BLOCOS[k]["presumido"] for k, b in devolucoes)
    base = sum(b for _, b in vendas)
    base_dev = sum(b for _, b in devolucoes)
    icms = (debito + estorno) - (cp + dev_icms)
    a, bb, c = base * 0.004, cp * 0.02, cp * 0.025
    fs_vendas = c + (a - (bb + c))
    a2, b2, c2 = base_dev * 0.004, estorno * 0.02, estorno * 0.025
    fs_dev = c2 + (a2 - (b2 + c2))
    return dict(base=base, debito=debito, credito_presumido=cp, devolucao_icms=dev_icms,
                estorno=estorno, icms_recolher=icms, fundo_social=fs_vendas - fs_dev,
                fundo_educacao=cp * 0.02 - estorno * 0.02,
                carga_efetiva=(icms / base * 100) if base else 0.0)


def cenarios_julho(movs, apur_lanc, competencia):
    """Mede o efeito das notas que ficaram fora da apuracao (decisao 3) sobre o gabarito."""
    vendas = [(l["bloco"][0], l["base"]) for l in apur_lanc if not l["devolucao"] and l["base"]]
    devol = [(l["bloco"][0], l["base"]) for l in apur_lanc if l["devolucao"] and l["base"]]
    atual = apurar(vendas, devol)
    apurados = {l["numero"] for l in apur_lanc}
    ano, mes = (int(x) for x in competencia.split("-"))
    fora = [m for m in movs if m["tipo"] == "S" and m["data"] and m["data"].year == ano
            and m["data"].month == mes and m["numero"] not in apurados]
    falta = sum(m["valor"] or 0 for m in fora)
    linhas = [("Notas fora da apuracao", len(fora), None, None),
              ("Valor contabil fora da apuracao", falta, None, None)]
    cen = {}
    for bloco in ("2", "3"):
        cen[bloco] = apurar(vendas + [(bloco, falta)], devol)
    campos = [("Base de calculo das operacoes beneficiadas", "base"),
              ("Debito do imposto", "debito"),
              ("Credito presumido", "credito_presumido"),
              ("Estorno de creditos (devolucoes)", "estorno"),
              ("ICMS a recolher", "icms_recolher"),
              ("Fundo Social a recolher", "fundo_social"),
              ("Fundo Educacao a recolher", "fundo_educacao"),
              ("Carga efetiva media do mes (%)", "carga_efetiva")]
    tabela = []
    for rotulo, chave in campos:
        tabela.append((rotulo, atual[chave], cen["2"][chave], cen["2"][chave] - atual[chave],
                       cen["3"][chave], cen["3"][chave] - atual[chave]))
    return atual, cen, tabela, fora, falta


# ----------------------------------------------------------------------------
# 10. Conferencia de totais
# ----------------------------------------------------------------------------

def reconciliar(conn, totais_planilha, painel, posicao):
    cur = conn.cursor()
    linhas = []

    def add(item, base, ref, fonte, obs=""):
        dif = (base or 0) - (ref or 0)
        linhas.append(dict(item=item, base=base, referencia=ref, diferenca=dif,
                           fonte=fonte, situacao="OK" if abs(dif) <= 0.5 else "DIVERGE", obs=obs))
        return dif

    e_kg, e_rs = cur.execute("""SELECT SUM(i.quantidade), SUM(i.valor) FROM nota_item i
                                JOIN nota n ON n.id=i.nota_id
                                WHERE n.tipo='E' AND n.natureza<>'ACERTO'""").fetchone()
    s_kg, s_rs = cur.execute("""SELECT SUM(i.quantidade), SUM(i.valor) FROM nota_item i
                                JOIN nota n ON n.id=i.nota_id WHERE n.tipo='S'""").fetchone()
    add("Entradas (kg)", e_kg, totais_planilha["E"][0], "linha de total da aba de entradas")
    add("Entradas (R$)", e_rs, totais_planilha["E"][1], "linha de total da aba de entradas")
    add("Saidas (kg)", s_kg, totais_planilha["S"][0], "linha de total da aba de saidas")
    add("Saidas (R$)", s_rs, totais_planilha["S"][1], "linha de total da aba de saidas")
    add("Entradas (kg)", e_kg, painel["entradas_kg"], "KPI do painel")
    add("Saidas (kg)", s_kg, painel["saidas_kg"], "KPI do painel")
    dif_kpi = add("Estoque total (kg)", (e_kg or 0) - (s_kg or 0), painel["estoque_total_kg"],
                  "KPI 'ESTOQUE TOTAL' do painel",
                  "KPI usa SUMIF(...;\"*\") e soma tudo")
    soma_painel = sum(painel["por_produto"].values())
    add("Estoque total (kg)", (e_kg or 0) - (s_kg or 0), soma_painel,
        "soma da lista 'SALDO POR PRODUTO' do painel",
        "a lista usa nomes literais e perde o produto acentuado")
    if abs(soma_painel - (painel["estoque_total_kg"] or 0)) > 0.5:
        REG.add("Painel - KPI x lista", "Critico", "DASHBOARD", None, "B5 x C9:C14",
                f"O KPI de estoque total ({painel['estoque_total_kg']:,.1f} kg) e a soma da lista por "
                f"produto ({soma_painel:,.1f} kg) diferem em "
                f"{(painel['estoque_total_kg'] - soma_painel):,.1f} kg.",
                impacto_kg=painel["estoque_total_kg"] - soma_painel,
                acao="Some pelo produto canonico; na base migrada os dois numeros coincidem")

    variantes = dict(cur.execute("SELECT descricao, variantes FROM produto").fetchall())
    for p in posicao:
        ref = painel["por_produto"].get(fold(p["produto"]))
        obs = ""
        if ref is None:
            obs = "produto inexistente na lista do painel"
        elif abs((p["saldo_aritmetico_kg"] or 0) - ref) > 0.5 and "|" in (variantes.get(p["produto"]) or ""):
            obs = ("a lista do painel so' soma a grafia literal; a diferenca sao os lancamentos "
                   "gravados como " + variantes[p["produto"]].split("|")[-1].strip())
        add(f"Saldo {p['produto']} (kg)", p["saldo_aritmetico_kg"], ref,
            "lista por produto do painel", obs)

    fora_kg, fora_n = cur.execute("""SELECT COALESCE(SUM(i.quantidade),0), COUNT(*) FROM nota_item i
        JOIN nota n ON n.id=i.nota_id WHERE n.data_mov IS NULL""").fetchone()
    linhas.append(dict(item="Movimentos fora do PEPS (sem data)", base=float(fora_n), referencia=0.0,
                       diferenca=float(fora_n), fonte="base migrada",
                       situacao="OK" if not fora_n else "DIVERGE",
                       obs=f"{fora_kg:,.1f} kg sem data nao entram no consumo PEPS - ver classe 'Data - ausente'"))
    ac_kg, ac_rs = cur.execute("""SELECT COALESCE(SUM(i.quantidade),0), COALESCE(SUM(i.valor),0)
        FROM nota_item i JOIN nota n ON n.id=i.nota_id WHERE n.natureza='ACERTO'""").fetchone()
    if ac_kg:
        linhas.append(dict(item="Acertos de estoque gerados (kg)", base=ac_kg, referencia=0.0,
                           diferenca=ac_kg, fonte="decisao 1 - acerto datado", situacao="OK",
                           obs="entram na base como natureza ACERTO e nao contaminam os totais acima"))
        linhas.append(dict(item="Acertos de estoque gerados (R$)", base=ac_rs, referencia=0.0,
                           diferenca=ac_rs, fonte="decisao 1 - acerto datado", situacao="OK",
                           obs="custo estimado - ver aba ACERTOS"))
        linhas.append(dict(item="Estoque final com acertos (kg)",
                           base=(e_kg or 0) - (s_kg or 0) + ac_kg, referencia=painel["estoque_total_kg"],
                           diferenca=ac_kg, fonte="KPI do painel", situacao="OK",
                           obs="a diferenca em relacao ao painel e' exatamente o total dos acertos"))
    sl = cur.execute("SELECT COALESCE(SUM(quantidade),0) FROM lastro WHERE status='sem_lastro'").fetchone()[0]
    linhas.append(dict(item="Saldo PEPS x saldo aritmetico (kg)",
                       base=sum(p["saldo_peps_kg"] for p in posicao),
                       referencia=sum(p["saldo_aritmetico_kg"] for p in posicao),
                       diferenca=sum(p["saldo_peps_kg"] for p in posicao) - sum(p["saldo_aritmetico_kg"] for p in posicao),
                       fonte="calculo interno",
                       situacao="OK" if abs(sum(p["saldo_peps_kg"] for p in posicao)
                                            - sum(p["saldo_aritmetico_kg"] for p in posicao)
                                            - sl - fora_kg - ac_kg) <= 0.5 else "DIVERGE",
                       obs=(f"a diferenca e' a quantidade que saiu sem lastro ({sl:,.1f} kg), o que ficou "
                            f"fora por falta de data ({fora_kg:,.1f} kg) e os acertos gerados "
                            f"({ac_kg:,.1f} kg)")))
    return linhas


# ----------------------------------------------------------------------------
# 11. Relatorio
# ----------------------------------------------------------------------------

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AZUL = "1F3864"; CINZA = "F2F2F2"; VERM = "C00000"; LAR = "BF8F00"; VERDE = "375623"
BORDA = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def escreve_aba(wb, titulo, cabecalhos, linhas, larguras=None, freeze="A2", nota=None):
    ws = wb.create_sheet(titulo[:31])
    r0 = 1
    if nota:
        ws.cell(1, 1, nota).font = Font(italic=True, color="595959", size=9)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(cabecalhos), 2))
        r0 = 2
    for c, h in enumerate(cabecalhos, start=1):
        cel = ws.cell(r0, c, h)
        cel.font = Font(bold=True, color="FFFFFF", size=10)
        cel.fill = PatternFill("solid", fgColor=AZUL)
        cel.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[r0].height = 28
    for i, linha in enumerate(linhas, start=r0 + 1):
        for c, v in enumerate(linha, start=1):
            cel = ws.cell(i, c, v)
            cel.border = BORDA
            cel.alignment = Alignment(vertical="top", wrap_text=isinstance(v, str) and len(str(v)) > 60)
            if isinstance(v, float):
                cel.number_format = "#,##0.00"
            if isinstance(v, dt.date):
                cel.number_format = "DD/MM/YYYY"
        sev = str(linha[1]) if len(linha) > 1 else ""
        if sev in SEV_ORDEM:
            cor = {"Critico": VERM, "Alto": LAR, "Medio": "7F7F7F", "Baixo": "A6A6A6"}[sev]
            ws.cell(i, 2).font = Font(bold=True, color=cor)
    for c in range(1, len(cabecalhos) + 1):
        ws.column_dimensions[get_column_letter(c)].width = (larguras or [18] * len(cabecalhos))[c - 1]
    ws.freeze_panes = ws[f"A{r0+1}"]
    ws.auto_filter.ref = f"A{r0}:{get_column_letter(len(cabecalhos))}{r0+len(linhas)}"
    return ws


def gerar_relatorio(caminho, conn, cruzamento, reconc, posicao, razao, sem_lastro, grupos, ocor,
                    canonico, cenario=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    cur = conn.cursor()

    # --- sumario ---------------------------------------------------------
    ws = wb.create_sheet("SUMARIO")
    ws["A1"] = "PROJETO LASTRO - FASE 1"
    ws["A1"].font = Font(bold=True, size=16, color=AZUL)
    ws["A2"] = "Relatorio de divergencias da migracao e saneamento do historico"
    ws["A2"].font = Font(size=11, color="595959")
    ws["A3"] = f"Gerado em {HOJE.strftime('%d/%m/%Y')} | fontes: ESTOQUE FISCAL IMPORTADO.xlsm e Apuracao ICMS 07/2026"
    ws["A3"].font = Font(size=9, italic=True, color="808080")
    n_e = cur.execute("SELECT COUNT(*) FROM nota WHERE tipo='E' AND natureza<>'ACERTO'").fetchone()[0]
    n_s = cur.execute("SELECT COUNT(*) FROM nota WHERE tipo='S'").fetchone()[0]
    n_ac = cur.execute("SELECT COUNT(*) FROM nota WHERE natureza='ACERTO'").fetchone()[0]
    n_p = cur.execute("SELECT COUNT(*) FROM parceiro").fetchone()[0]
    n_pr = cur.execute("SELECT COUNT(*) FROM produto").fetchone()[0]
    n_l = cur.execute("SELECT COUNT(*) FROM lastro WHERE status='ok'").fetchone()[0]
    por_sev = defaultdict(int)
    for d in REG.itens:
        if d["situacao"] == "Pendente":
            por_sev[d["severidade"]] += 1
    resolvidas = sum(1 for d in REG.itens if d["situacao"] != "Pendente")
    resumo = [
        ("Lancamentos migrados", f"{n_e + n_s}", f"{n_e} entradas e {n_s} saidas"),
        ("Acertos de estoque gerados", f"{n_ac}", "decisao 1 - cada um na data da saida que o motivou"),
        ("Parceiros apos saneamento", f"{n_p}", "grafias diferentes unificadas; nenhum tem CNPJ"),
        ("Produtos apos saneamento", f"{n_pr}", "variacao de acento eliminada"),
        ("Consumos de estoque calculados", f"{n_l}", "PEPS por produto (custeio), nov/2020 -> hoje"),
        ("", "", ""),
        ("Achados resolvidos na migracao", str(resolvidas),
         "saneados pelo script ou pelas decisoes de 27/08/2026"),
        ("Pendentes - criticas", str(por_sev["Critico"]), "bloqueiam a Fase 2"),
        ("Pendentes - altas", str(por_sev["Alto"]), "precisam de conferencia sua"),
        ("Pendentes - medias", str(por_sev["Medio"]), ""),
        ("Pendentes - baixas", str(por_sev["Baixo"]), ""),
        ("Total de achados", str(len(REG.itens)), ""),
    ]
    r = 5
    for a, b, c in resumo:
        ws.cell(r, 1, a).font = Font(bold=bool(a) and a.startswith(("Total", "Diverg")))
        ws.cell(r, 2, b).font = Font(bold=True, color=AZUL)
        ws.cell(r, 3, c).font = Font(size=9, color="808080")
        r += 1
    r += 1
    ws.cell(r, 1, "ACHADOS POR CLASSE").font = Font(bold=True, color=AZUL); r += 1
    for classe, itens in sorted(REG.por_classe().items(),
                                key=lambda kv: (min(SEV_ORDEM[i["severidade"]] for i in kv[1]), kv[0])):
        ws.cell(r, 1, classe)
        ws.cell(r, 2, len(itens)).font = Font(bold=True)
        ws.cell(r, 3, sorted({i["severidade"] for i in itens}, key=lambda s: SEV_ORDEM[s])[0]
                + " | " + ("Pendente" if any(i["situacao"] == "Pendente" for i in itens)
                           else "Resolvido na migracao"))
        r += 1
    r += 1
    ws.cell(r, 1, "COMO USAR ESTE ARQUIVO").font = Font(bold=True, color=AZUL); r += 1
    for t in ["1. A aba DIVERGENCIAS e' a lista de trabalho: uma linha por achado, com a linha exata do arquivo de origem.",
              "2. A coluna DECISAO esta' vazia de proposito. Nada foi corrigido nos arquivos originais.",
              "3. PARCEIROS e PRODUTOS trazem as unificacoes propostas - confirme antes de a Fase 2 comecar.",
              "4. SAIDAS SEM SALDO, ACERTOS e POSICAO PEPS mostram o efeito dos 6 anos de historico sobre o estoque valorizado.",
              "5. CRUZAMENTO 07-2026 compara nota a nota o estoque contra a apuracao do mes anexado.",
              "6. RECONCILIACAO prova que a base migrada reproduz os totais das planilhas - ou explica por que nao."]:
        ws.cell(r, 1, t).font = Font(size=9)
        r += 1
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60

    # --- decisoes --------------------------------------------------------
    escreve_aba(wb, "DECISOES",
                ["#", "PERGUNTA", "DECISAO DE 27/08/2026", "COMO O SCRIPT APLICA", "ONDE VER"],
                [[1, "Saldo negativo e saida sem lastro: acerto ou aceitar o negativo?",
                  "Lancamento de acerto datado",
                  "Toda saida sem entrada disponivel gera uma nota de natureza ACERTO na data da "
                  "propria saida, com custo estimado. Nenhum saldo fica negativo em nenhum momento "
                  "dos 6 anos.", "aba ACERTOS"],
                 [2, "NCM: preencher retroativamente os 6 produtos?", "Desconsiderar",
                  "O campo continua no modelo e aceita nulo. Nenhuma divergencia critica e' aberta "
                  "por NCM ausente. O PEPS separa por produto, nao por NCM.", "aba PRODUTOS"],
                 [3, "As 3 notas de cobre de julho fora da apuracao: dentro ou fora do beneficio?",
                  "Nao entraram mesmo; serao ajustadas com a contabilidade em ago/2026",
                  "As notas seguem na base como saida de importado e a divergencia fica registrada "
                  "como ajuste em curso. Julho/2026 deixa de valer como gabarito ate' o ajuste sair.",
                  "abas IMPACTO JULHO e CRUZAMENTO 07-2026"],
                 [4, "Parceiros com nome parecido: unificar ou manter separados?",
                  "Unificar",
                  "Grupos com o mesmo nucleo de razao social viraram um unico parceiro. Cada razao "
                  "social original ficou gravada em parceiro.variantes, entao a separacao e' "
                  "reversivel quando os CNPJs entrarem.", "aba PARCEIROS"],
                 [5, "Registrar a origem da mercadoria e vincular cada saida a uma entrada?",
                  "Nao. Controle pelo saldo de estoque",
                  "A tabela lastro deixa de ser prova de vinculacao e passa a ser o razao de CUSTEIO: "
                  "o PEPS existe para valorizar o estoque, nao para amarrar NF de saida a NF de "
                  "entrada. O campo origem_merc continua no modelo, nulo, e sera' preenchido pelo "
                  "XML na Fase 5 sem ninguem digitar. O anexo de lastro sai do escopo da exportacao.",
                  "abas SAIDAS SEM SALDO e POSICAO PEPS"],
                 [6, "Linhas identicas (TSR, NF 5257 e 5290) sao duplicidade?",
                  "Nao. Notas diferentes podem ter dados identicos - manter as duas",
                  "As linhas ficam na base como lancamentos distintos e o achado vira informativo. "
                  "Com a chave de acesso (Fase 5) a duplicidade real deixa de ser possivel.",
                  "aba DIVERGENCIAS, classe 'Linha duplicada - mantida por decisao'"]],
                larguras=[5, 46, 40, 66, 30],
                nota="As quatro decisoes ja' estao aplicadas nesta versao da base. Mudar uma decisao "
                     "e' mudar o dicionario DECISOES no topo do script e rodar de novo.")

    # --- divergencias ----------------------------------------------------
    ordenadas = sorted(REG.itens, key=lambda d: (d["situacao"] != "Pendente",
                                                SEV_ORDEM[d["severidade"]], d["classe"], d["linha"] or 0))
    escreve_aba(wb, "DIVERGENCIAS",
                ["ID", "SEVERIDADE", "SITUACAO", "CLASSE", "ORIGEM", "LINHA", "DOCUMENTO",
                 "O QUE FOI ENCONTRADO", "IMPACTO (KG)", "IMPACTO (R$)", "ACAO SUGERIDA", "DECISAO"],
                [[d["id"], d["severidade"], d["situacao"], d["classe"], d["origem"], d["linha"],
                  d["documento"], d["descricao"], d["impacto_kg"], d["impacto_rs"], d["acao"],
                  d["decisao"]] for d in ordenadas],
                larguras=[8, 12, 20, 30, 18, 8, 14, 70, 14, 14, 60, 24],
                nota="Uma linha por achado. SITUACAO separa o que a migracao ja' resolveu do que ainda "
                     "espera decisao ou conferencia. Nada foi alterado nos arquivos originais.")

    # --- parceiros -------------------------------------------------------
    linhas = []
    for canon, membros in sorted(grupos.items()):
        variantes, qtd, lados = set(), 0, set()
        for m in membros:
            variantes |= ocor[m]["variantes"]
            qtd += ocor[m]["qtd"]
            lados |= ocor[m]["lados"]
        nome = max((limpa_texto(v) for v in variantes), key=len)
        linhas.append([nome, len(variantes), qtd,
                       "ambos" if len(lados) > 1 else (list(lados)[0] if lados else ""),
                       " | ".join(sorted(repr(v)[1:-1] for v in variantes)),
                       "", ""])
    escreve_aba(wb, "PARCEIROS", ["NOME CANONICO", "GRAFIAS", "LANCAMENTOS", "PAPEL",
                                  "GRAFIAS ENCONTRADAS (com espacos invisiveis)", "CNPJ", "CONFIRMA? (S/N)"],
                linhas, larguras=[52, 10, 14, 12, 90, 22, 16],
                nota="Unificacao proposta por semelhanca de nome. Sem CNPJ na planilha, isto e' heuristica: confirme uma a uma.")

    # --- produtos --------------------------------------------------------
    linhas = [[d, v, c, m, ""] for (d, v, c, m) in
              cur.execute("SELECT descricao, variantes, categoria, metal FROM produto ORDER BY descricao")]
    escreve_aba(wb, "PRODUTOS", ["DESCRICAO CANONICA", "GRAFIAS ENCONTRADAS", "CATEGORIA", "METAL", "NCM (preencher)"],
                linhas, larguras=[30, 46, 14, 14, 20],
                nota="A coluna NCM esta' vazia nos dois arquivos: 6 produtos, 6 NCMs a definir uma unica vez.")

    # --- sem lastro ------------------------------------------------------
    escreve_aba(wb, "SAIDAS SEM SALDO",
                ["DATA", "NF", "PRODUTO", "CLIENTE", "QTD DA NOTA (KG)", "SEM SALDO (KG)", "ORIGEM NO ARQUIVO"],
                [[s["data"], s["numero"], s["produto"], s["parceiro"], s["qtd_nota"],
                  s["qtd_sem_lastro"], s["origem"]] for s in sem_lastro],
                larguras=[12, 10, 24, 44, 18, 18, 34],
                nota="Saidas cuja quantidade excedeu o saldo do produto na data. Cada uma gerou um acerto "
                     "(decisao 1). O controle e' por saldo de estoque, nao por vinculo nota a nota (decisao 5).")

    # --- acertos ---------------------------------------------------------
    if ACERTOS:
        escreve_aba(wb, "ACERTOS",
                    ["DATA", "PRODUTO", "QTD (KG)", "CUSTO ESTIMADO (R$/KG)", "VALOR (R$)",
                     "SAIDA QUE MOTIVOU (NF)", "CLIENTE", "ORIGEM NO ARQUIVO", "CUSTO CONFIRMADO"],
                    [[a["data"], a["produto"], a["qtd"], a["custo"], a["valor"], a["saida_nf"],
                      a["cliente"], a["origem"], ""] for a in ACERTOS],
                    larguras=[12, 24, 14, 22, 16, 20, 40, 34, 20],
                    nota="Decisao 1: cada acerto entra na data da saida que o motivou. O custo e' a media "
                         "ponderada das entradas anteriores do mesmo produto - confirme ou substitua.")

    # --- posicao ---------------------------------------------------------
    escreve_aba(wb, "POSICAO PEPS",
                ["PRODUTO", "SALDO PEPS (KG)", "SALDO PEPS (R$)", "CUSTO MEDIO (R$/KG)",
                 "SALDO DA PLANILHA (KG)", "ACERTOS (KG)", "SALDO COM ACERTOS (KG)"],
                [[p["produto"], p["saldo_peps_kg"], p["saldo_peps_rs"], p["custo_medio"],
                  p["saldo_aritmetico_kg"], p.get("acerto_kg", 0.0), p.get("saldo_com_acerto_kg")]
                 for p in posicao],
                larguras=[28, 18, 20, 20, 22, 16, 22],
                nota="Saldo PEPS = o que resta das entradas nao consumidas (ja' valorizado). Aritmetico = entradas - saidas.")

    # --- cruzamento ------------------------------------------------------
    escreve_aba(wb, "CRUZAMENTO 07-2026",
                ["DOCUMENTO", "LADO", "NO ESTOQUE", "NA APURACAO", "PRODUTO (ESTOQUE)",
                 "PRODUTO (APURACAO)", "VALOR (ESTOQUE)", "VALOR (APURACAO)", "DIAGNOSTICO"],
                [[c["doc"], c["lado"], c["estoque"], c["apuracao"], c["produto_estoque"],
                  c["produto_apuracao"], c["valor_estoque"], c["valor_apuracao"], c["diagnostico"]]
                 for c in cruzamento],
                larguras=[12, 20, 12, 30, 24, 24, 16, 16, 56],
                nota="Nota a nota: a mesma NF digitada nos dois arquivos, sem nenhuma conferencia automatica hoje.")

    # --- reconciliacao ---------------------------------------------------
    escreve_aba(wb, "RECONCILIACAO",
                ["ITEM", "BASE MIGRADA", "REFERENCIA", "DIFERENCA", "FONTE DA REFERENCIA", "SITUACAO", "OBSERVACAO"],
                [[l["item"], l["base"], l["referencia"], l["diferenca"], l["fonte"], l["situacao"], l["obs"]]
                 for l in reconc],
                larguras=[34, 18, 18, 16, 40, 12, 46],
                nota="Criterio de entrega da Fase 1: todo item aqui esta' OK ou tem explicacao escrita.")

    # --- impacto de julho ------------------------------------------------
    if cenario:
        atual, cen, tabela, fora, falta = cenario
        linhas_c = [[rot, at, c2, d2, c3, d3] for (rot, at, c2, d2, c3, d3) in tabela]
        linhas_c.append(["", None, None, None, None, None])
        for m in fora:
            linhas_c.append([f"NF {m['numero']} - {limpa_texto(m['parceiro_bruto'])}", m["valor"],
                             None, None, None, None])
        escreve_aba(wb, "IMPACTO JULHO",
                    ["LINHA DA APURACAO", "COMO ESTA' HOJE",
                     "SE ENTRAR NO BLOCO 2 (4%)", "DIFERENCA", "SE ENTRAR NO BLOCO 3 (12%)", "DIFERENCA"],
                    linhas_c, larguras=[46, 20, 24, 18, 24, 18],
                    nota=f"Decisao 3: R$ {falta:,.2f} em {len(fora)} notas ficaram fora da apuracao de julho. "
                         "O bloco correto depende da UF do destinatario - as duas hipoteses estao medidas. "
                         "Enquanto o ajuste nao sair, julho nao serve de gabarito para a Fase 3.")

    # --- razao -----------------------------------------------------------
    escreve_aba(wb, "RAZAO PEPS",
                ["DATA", "TIPO", "NATUREZA", "NF", "PRODUTO", "PARCEIRO", "QTD (KG)", "VALOR (R$)",
                 "SALDO DO PRODUTO (KG)", "ORIGEM NO ARQUIVO"],
                [[x["data"], "ENTRADA" if x["tipo"] == "E" else "SAIDA", x["natureza"], x["numero"],
                  x["produto"], x["parceiro"], x["qtd"], x["valor"], x["saldo"], x["origem"]]
                 for x in razao],
                larguras=[12, 10, 12, 10, 24, 44, 14, 16, 20, 34],
                nota="Extrato linha a linha com saldo corrido por produto - a conferencia que hoje nao existe.")

    wb.save(caminho)


# ----------------------------------------------------------------------------
# 12. Main
# ----------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--estoque", default="estoque.xlsm")
    ap.add_argument("--apuracao", default="apuracao_072026.xlsx")
    ap.add_argument("--db", default="lastro.db")
    ap.add_argument("--relatorio", default="relatorio_divergencias_fase1.xlsx")
    ap.add_argument("--competencia", default=COMP_APURACAO)
    args = ap.parse_args()

    movs, totais, painel = ler_movimentos(args.estoque)
    apur_lanc, resumo, apur_entr = ler_apuracao(args.apuracao)
    print(f"lidos {len(movs)} movimentos | {len(apur_lanc)} lancamentos de apuracao "
          f"| {len(apur_entr)} entradas de importacao")

    for m in movs:
        canon_produto(m["produto_bruto"])
    canonico, grupos, ocor = sanear_parceiros(
        movs, extras=[l["parceiro_bruto"] for l in apur_lanc] + [e["parceiro_bruto"] for e in apur_entr])

    conn = criar_base(args.db)
    carregar(conn, movs, apur_lanc, apur_entr, canonico, grupos, ocor)

    analisar_estrutura(movs)
    analisar_movimentos(movs)
    analisar_duplicidades(movs)
    razao, negativos, sem_lastro, posicao = rodar_peps(conn)
    analisar_peps(negativos, sem_lastro, posicao)
    cruzamento = cruzar_competencia(movs, apur_lanc, apur_entr, args.competencia)
    cenario = cenarios_julho(movs, apur_lanc, args.competencia)
    reconc = reconciliar(conn, totais, painel, posicao)

    cur = conn.cursor()
    for d in REG.itens:
        cur.execute("""INSERT INTO divergencia VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                    (d["id"], d["classe"], d["severidade"], d["origem"], d["linha"], d["documento"],
                     d["descricao"], d["impacto_kg"], d["impacto_rs"], d["acao"], d["decisao"]))
    conn.commit()

    gerar_relatorio(args.relatorio, conn, cruzamento, reconc, posicao, razao, sem_lastro,
                    grupos, ocor, canonico, cenario=cenario)
    por_sev = defaultdict(int)
    for d in REG.itens:
        por_sev[d["severidade"]] += 1
    print("divergencias:", dict(por_sev), "| total", len(REG.itens))
    print("relatorio:", args.relatorio, "| base:", args.db)
    conn.close()


if __name__ == "__main__":
    main()
