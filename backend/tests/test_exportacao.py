"""O arquivo gerado tem que ser o mesmo que a contabilidade ja' recebe."""
import openpyxl
import pytest

from app.services.exportacao import exportar_apuracao, exportar_estoque, exportar_pendencias

GABARITO_BLOCOS = {"2": 2045095.20, "3": 1455407.20}


@pytest.fixture(scope="module")
def planilha(request):
    return None


def _ws(db, comp="2026-07"):
    return openpyxl.load_workbook(exportar_apuracao(db, comp))


def test_estrutura_da_apuracao(db):
    wb = _ws(db)
    assert wb.sheetnames == ["OPERAÇÕES SAÍDA IMPORTADO", "OPERAÇÕES ENTRADA IMPORTADO",
                             "ORIGEM DO ARQUIVO"]
    ws = wb["OPERAÇÕES SAÍDA IMPORTADO"]
    assert ws["A1"].value.startswith("SAÍDAS COM BENEFÍCIO FISCAL")
    textos = [c.value for c in ws["A"] if isinstance(c.value, str)]
    for esperado in ["DEVOLUÇÃO DE VENDAS COM BENEFÍCIO FISCAL – TTD 409, 410 OU 411",
                     "CRÉDITOS", "DÉBITOS", "FUNDO SOCIAL", "DEVOLUÇÕES", "TOTAL A RECOLHER"]:
        assert any(t.startswith(esperado[:20]) for t in textos), esperado


def test_lancamentos_e_totais_por_bloco(db):
    ws = _ws(db)["OPERAÇÕES SAÍDA IMPORTADO"]
    docs = [c.value for c in ws["B"] if isinstance(c.value, int) and c.value > 1000]
    for nf in (6448, 6466, 6481, 6524, 6534, 6456, 6477, 6491, 6494, 6498, 6535, 1459):
        assert nf in docs, nf
    # Seis blocos (3 de saida + 3 de devolucao), cada um com sua linha de TOTAL.
    totais = [c.row for c in ws["A"] if c.value == "TOTAL"]
    assert len(totais) == 6
    # Onde ha' lancamento, o total e' formula sobre o intervalo EXATO do bloco - nada de faixa
    # fixa de 15 linhas somando o que nao e' dele.
    formulas = [c.value for c in ws["E"] if isinstance(c.value, str) and c.value.startswith("=SUM")]
    assert formulas == ["=SUM(E8:E12)", "=SUM(E17:E22)", "=SUM(E32:E32)"]
    # Bloco vazio nao vira formula sobre intervalo invertido: vai zero.
    assert [ws.cell(r, 5).value for r in totais if ws.cell(r, 5).value == 0] == [0, 0, 0]


def test_resumo_de_sete_linhas(db):
    ws = _ws(db)["OPERAÇÕES SAÍDA IMPORTADO"]
    numeradas = [(c.row, c.value) for c in ws["A"] if isinstance(c.value, int) and 1 <= c.value <= 7]
    assert [v for _, v in numeradas] == [1, 2, 3, 4, 5, 6, 7]
    valores = {v: ws.cell(r, 7).value for r, v in numeradas}
    assert abs(valores[1] - 256452.672) < 0.01
    assert abs(valores[4] - 205438.1688) < 0.01
    assert abs(valores[5] - 44914.3762) < 0.01
    assert abs(valores[6] - 7819.203044) < 0.01
    assert abs(valores[7] - 3742.755756) < 0.01


def test_entradas_de_importacao_do_mes(db):
    ws = _ws(db)["OPERAÇÕES ENTRADA IMPORTADO"]
    numeros = [ws.cell(r, 3).value for r in range(2, ws.max_row + 1)]
    for nf in (6467, 6492, 6509):
        assert nf in numeros, nf


def test_exportacao_de_estoque(db):
    wb = openpyxl.load_workbook(exportar_estoque(db, None, None))
    assert wb.sheetnames == ["POSIÇÃO", "MOVIMENTAÇÃO"]
    ws = wb["POSIÇÃO"]
    produtos = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
    assert "SUCATA DE ALUMINIO" in produtos and "TOTAL" in produtos
    mov = wb["MOVIMENTAÇÃO"]
    assert mov.max_row > 900        # os 6 anos, linha a linha


def test_relatorio_de_erros_agrupa_por_area(db):
    """O historico migrado ja' tem pendencia de estoque (acerto automatico) e de importacao
    (nota sem data) - o relatorio precisa separar isso, nao virar lista solta."""
    wb = openpyxl.load_workbook(exportar_pendencias(db))
    assert wb.sheetnames == ["RESUMO", "PENDÊNCIAS"]
    resumo = [c.value for c in wb["RESUMO"]["A"] if isinstance(c.value, str)]
    assert "Estoque" in resumo or any("Estoque" in (t or "") for t in resumo)
    ws = wb["PENDÊNCIAS"]
    areas = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}
    assert "Estoque" in areas and "Importação" in areas
